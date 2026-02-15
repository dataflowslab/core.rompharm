"""
Routes pentru modulul Buget Aprobat
Gestionează operațiunile CRUD pentru bugetul aprobat
"""
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from typing import Optional, List, Dict, Any
from datetime import datetime, date
from bson import ObjectId, Decimal128
from decimal import Decimal
import re
import io

from ..utils.db import get_db
from ..routes.auth import verify_token

router = APIRouter(prefix="/api/buget", tags=["buget"])

# ============================================================================
# HELPER FUNCTIONS
# ============================================================================

def decimal128_to_float(value):
    """Convertește Decimal128 în float"""
    if isinstance(value, Decimal128):
        return float(value.to_decimal())
    return value if value else 0


def parseaza_cod_clasificare(cod: str) -> dict:
    """Parsează codul de clasificare bugetară"""
    if not cod or len(cod) < 10:
        return {"capitol": "", "subcapitol": "", "articol": "", "alineat": ""}
    return {
        "capitol": cod[0:4],
        "subcapitol": cod[4:8],
        "articol": cod[8:10],
        "alineat": cod[10:] if len(cod) > 10 else ""
    }

def format_buget_item(item: dict) -> dict:
    """Formatează un item de buget pentru răspuns"""
    return {
        '_id': str(item['_id']),
        'cod_clasificare': item.get('cod_clasificare', ''),
        'cod_clasificare_parsed': item.get('cod_clasificare_parsed', {}),
        'denumire': item.get('denumire', ''),
        'cod_subunitate': item.get('cod_subunitate', ''),
        'cod_program': item.get('cod_program', ''),
        'tip_inregistrare': item.get('tip_inregistrare', ''),
        'tip_operatie': item.get('tip_operatie', ''),
        
        # Sume
        'suma_initiala': decimal128_to_float(item.get('suma_initiala')),
        'suma_curenta': decimal128_to_float(item.get('suma_curenta')),
        'trimestre': {
            't1': decimal128_to_float(item.get('trimestre', {}).get('t1')),
            't2': decimal128_to_float(item.get('trimestre', {}).get('t2')),
            't3': decimal128_to_float(item.get('trimestre', {}).get('t3')),
            't4': decimal128_to_float(item.get('trimestre', {}).get('t4'))
        },
        
        # Document
        'numar_document': item.get('numar_document', ''),
        'data_document': item.get('data_document').isoformat() if item.get('data_document') else None,
        
        # Referințe
        'capitol_id': str(item['capitol_id']) if item.get('capitol_id') else None,
        'subcapitol_id': str(item['subcapitol_id']) if item.get('subcapitol_id') else None,
        'articol_id': str(item['articol_id']) if item.get('articol_id') else None,
        'alineat_id': str(item['alineat_id']) if item.get('alineat_id') else None,
        
        # Metadata
        'an_bugetar': item.get('an_bugetar', 2024),
        'activ': item.get('activ', True),
        'data_actualizare': item.get('data_actualizare').isoformat() if item.get('data_actualizare') else None,
        'utilizator_id': item.get('utilizator_id', ''),
        'created_at': item.get('created_at').isoformat() if item.get('created_at') else None,
        'updated_at': item.get('updated_at').isoformat() if item.get('updated_at') else None
    }


def _parse_decimal(value: Any) -> Decimal:
    try:
        if value is None or value == '':
            return Decimal(0)
        if isinstance(value, Decimal):
            return value
        return Decimal(str(value).replace(',', '.'))
    except Exception:
        return Decimal(0)

# ============================================================================
# BUGET APROBAT - LIST & SEARCH
# ============================================================================

@router.get("/initial")
async def get_buget_initial_list(
    page: int = Query(1, ge=1),
    limit: int = Query(50, ge=1, le=500),
    search: Optional[str] = None,
    an_bugetar: Optional[int] = None,
    user = Depends(verify_token)
) -> Dict[str, Any]:
    """
    Obține lista bugetului inițial (read-only)
    """
    db = get_db()
    query = {}
    if search:
        search_regex = re.compile(re.escape(search), re.IGNORECASE)
        query['$or'] = [
            {'cod_clasificare': search_regex},
            {'denumire': search_regex}
        ]
    if an_bugetar:
        query['an_bugetar'] = an_bugetar
    skip = (page - 1) * limit
    cursor = db.alop_buget_initial.find(query).sort([('cod_clasificare', 1)]).skip(skip).limit(limit)
    items = [format_buget_item(item) for item in cursor]
    total = db.alop_buget_initial.count_documents(query)
    return {
        'items': items,
        'total': total,
        'page': page,
        'limit': limit,
        'pages': (total + limit - 1) // limit
    }


@router.get("/initial/exists")
async def buget_initial_exists(
    an_bugetar: int = Query(...),
    user = Depends(verify_token)
) -> Dict[str, Any]:
    """Check if initial budget already loaded for given year."""
    db = get_db()
    count = db.alop_buget_initial.count_documents({'an_bugetar': an_bugetar})
    return {'exists': count > 0, 'count': count}


@router.get("/initial/template")
async def download_buget_initial_template(
    user = Depends(verify_token)
):
    """Download template for buget initial import."""
    from fastapi.responses import StreamingResponse
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        raise HTTPException(status_code=500, detail="Modulul openpyxl nu este instalat")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Buget Initial"
    headers = [
        "Cod Clasificare",
        "Denumire",
        "Suma Inițială",
        "Trimestrul 1",
        "Trimestrul 2",
        "Trimestrul 3",
        "Trimestrul 4",
        "Subunitate",
        "Program"
    ]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    # Example row
    ws.cell(row=2, column=1, value="65000102101")
    ws.cell(row=2, column=2, value="Exemplu denumire")
    ws.cell(row=2, column=3, value=10000)
    ws.cell(row=2, column=4, value=2500)
    ws.cell(row=2, column=5, value=2500)
    ws.cell(row=2, column=6, value=2500)
    ws.cell(row=2, column=7, value=2500)
    ws.cell(row=2, column=8, value="0001")
    ws.cell(row=2, column=9, value="01")

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"buget_initial_template_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.post("/initial/upload")
async def upload_buget_initial(
    an_bugetar: int = Query(...),
    file: UploadFile = File(...),
    user = Depends(verify_token)
) -> Dict[str, Any]:
    """Upload buget initial from xlsx. One upload per year."""
    db = get_db()

    if db.alop_buget_initial.count_documents({'an_bugetar': an_bugetar}) > 0:
        raise HTTPException(status_code=400, detail="Bugetul inițial pentru acest an este deja încărcat")

    if not file.filename or not file.filename.lower().endswith('.xlsx'):
        raise HTTPException(status_code=400, detail="Fișierul trebuie să fie .xlsx")

    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="Modulul openpyxl nu este instalat")

    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active

    docs = []
    now = datetime.utcnow()
    row_index = 2
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or all(cell is None or str(cell).strip() == '' for cell in row):
            row_index += 1
            continue
        cod_clasificare = str(row[0]).strip() if len(row) > 0 and row[0] is not None else ''
        denumire = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ''
        suma_initiala = _parse_decimal(row[2] if len(row) > 2 else 0)
        t1 = _parse_decimal(row[3] if len(row) > 3 else 0)
        t2 = _parse_decimal(row[4] if len(row) > 4 else 0)
        t3 = _parse_decimal(row[5] if len(row) > 5 else 0)
        t4 = _parse_decimal(row[6] if len(row) > 6 else 0)
        cod_subunitate = str(row[7]).strip() if len(row) > 7 and row[7] is not None else ''
        cod_program = str(row[8]).strip() if len(row) > 8 and row[8] is not None else ''

        if not cod_clasificare:
            row_index += 1
            continue

        doc = {
            'cod_clasificare': cod_clasificare,
            'cod_clasificare_parsed': parseaza_cod_clasificare(cod_clasificare),
            'denumire': denumire,
            'cod_subunitate': cod_subunitate,
            'cod_program': cod_program,
            'tip_inregistrare': 'INITIAL',
            'tip_operatie': 'INITIAL',
            'suma_initiala': Decimal128(suma_initiala),
            'suma_curenta': Decimal128(suma_initiala),
            'trimestre': {
                't1': Decimal128(t1),
                't2': Decimal128(t2),
                't3': Decimal128(t3),
                't4': Decimal128(t4)
            },
            'numar_document': '',
            'data_document': None,
            'an_bugetar': an_bugetar,
            'activ': True,
            'data_actualizare': now,
            'utilizator_id': user.get('uid', ''),
            'created_at': now,
            'updated_at': now
        }
        docs.append(doc)
        row_index += 1

    if not docs:
        raise HTTPException(status_code=400, detail="Fișierul nu conține date valide")

    db.alop_buget_initial.insert_many(docs)

    # Seed buget aprobat if empty for year
    inserted_aprobat = 0
    if db.alop_buget_aprobat.count_documents({'an_bugetar': an_bugetar}) == 0:
        aprobat_docs = []
        for doc in docs:
            clone = dict(doc)
            clone.pop('_id', None)
            aprobat_docs.append(clone)
        if aprobat_docs:
            db.alop_buget_aprobat.insert_many(aprobat_docs)
            inserted_aprobat = len(aprobat_docs)

    return {
        'success': True,
        'count': len(docs),
        'seeded_buget_aprobat': inserted_aprobat
    }

@router.get("/aprobat")
async def get_buget_aprobat_list(
    page: int = Query(1, ge=1),
    limit: int = Query(50, ge=1, le=500),
    search: Optional[str] = None,
    capitol: Optional[str] = None,
    subcapitol: Optional[str] = None,
    an_bugetar: Optional[int] = None,
    tip_operatie: Optional[str] = None,
    user = Depends(verify_token)
) -> Dict[str, Any]:
    """
    Obține lista bugetului aprobat cu paginare și căutare AJAX
    """
    db = get_db()
    
    # Construire query
    query = {}
    
    if search:
        # Căutare în cod_clasificare și denumire
        search_regex = re.compile(re.escape(search), re.IGNORECASE)
        query['$or'] = [
            {'cod_clasificare': search_regex},
            {'denumire': search_regex}
        ]
    
    if capitol:
        query['cod_clasificare_parsed.capitol'] = capitol
    
    if subcapitol:
        query['cod_clasificare_parsed.subcapitol'] = subcapitol
    
    if an_bugetar:
        query['an_bugetar'] = an_bugetar
    
    if tip_operatie:
        query['tip_operatie'] = tip_operatie
    
    # Calculare skip
    skip = (page - 1) * limit
    
    # Obținere date
    cursor = db.alop_buget_aprobat.find(query).sort([
        ('cod_clasificare', 1)
    ]).skip(skip).limit(limit)
    
    items = []
    for item in cursor:
        items.append(format_buget_item(item))
    
    # Total
    total = db.alop_buget_aprobat.count_documents(query)
    
    return {
        'items': items,
        'total': total,
        'page': page,
        'limit': limit,
        'pages': (total + limit - 1) // limit
    }

@router.get("/aprobat/search")
async def search_buget_aprobat(
    q: str = Query(..., min_length=1),
    limit: int = Query(20, ge=1, le=100),
    user = Depends(verify_token)
) -> List[Dict[str, Any]]:
    """
    Căutare AJAX pentru buget aprobat
    """
    db = get_db()
    
    search_regex = re.compile(re.escape(q), re.IGNORECASE)
    
    cursor = db.alop_buget_aprobat.find({
        '$or': [
            {'cod_clasificare': search_regex},
            {'denumire': search_regex}
        ]
    }).limit(limit)
    
    results = []
    for item in cursor:
        results.append({
            '_id': str(item['_id']),
            'cod_clasificare': item.get('cod_clasificare', ''),
            'denumire': item.get('denumire', ''),
            'suma_curenta': decimal128_to_float(item.get('suma_curenta')),
            'label': f"{item.get('cod_clasificare', '')} - {item.get('denumire', '')}"
        })
    
    return results

@router.get("/aprobat/{item_id}")
async def get_buget_aprobat_item(
    item_id: str,
    user = Depends(verify_token)
) -> Dict[str, Any]:
    """
    Obține un item de buget aprobat după ID
    """
    db = get_db()
    
    try:
        item = db.alop_buget_aprobat.find_one({'_id': ObjectId(item_id)})
    except:
        raise HTTPException(status_code=400, detail="ID invalid")
    
    if not item:
        raise HTTPException(status_code=404, detail="Înregistrare negăsită")
    
    return format_buget_item(item)

# ============================================================================
# BUGET APROBAT - MODIFICARE
# ============================================================================

@router.put("/aprobat/{item_id}")
async def update_buget_aprobat(
    item_id: str,
    data: Dict[str, Any],
    user = Depends(verify_token)
) -> Dict[str, Any]:
    """
    Modifică un item de buget aprobat
    Nu suprascrie bugetul inițial, ci creează o nouă înregistrare în alop_buget_aprobat
    """
    db = get_db()
    
    try:
        existing = db.alop_buget_aprobat.find_one({'_id': ObjectId(item_id)})
    except:
        raise HTTPException(status_code=400, detail="ID invalid")
    
    if not existing:
        raise HTTPException(status_code=404, detail="Înregistrare negăsită")
    
    # Validare câmpuri obligatorii
    if not data.get('numar_document'):
        raise HTTPException(status_code=400, detail="Numărul documentului este obligatoriu")
    
    if not data.get('data_document'):
        raise HTTPException(status_code=400, detail="Data documentului este obligatorie")
    
    # Calculare suma totală din trimestre dacă sunt completate
    t1 = Decimal(str(data.get('trimestru_1', 0) or 0))
    t2 = Decimal(str(data.get('trimestru_2', 0) or 0))
    t3 = Decimal(str(data.get('trimestru_3', 0) or 0))
    t4 = Decimal(str(data.get('trimestru_4', 0) or 0))
    
    suma_trimestre = t1 + t2 + t3 + t4
    suma_totala = Decimal(str(data.get('suma_totala', 0) or 0))
    
    # Dacă suma trimestrelor > 0, folosește-o
    if suma_trimestre > 0:
        suma_totala = suma_trimestre
    
    # Pregătire update
    update_data = {
        'suma_curenta': Decimal128(suma_totala),
        'trimestre': {
            't1': Decimal128(t1),
            't2': Decimal128(t2),
            't3': Decimal128(t3),
            't4': Decimal128(t4)
        },
        'numar_document': data.get('numar_document'),
        'data_document': datetime.fromisoformat(data['data_document']) if isinstance(data.get('data_document'), str) else data.get('data_document'),
        'tip_operatie': 'MODIFICARE',
        'updated_at': datetime.utcnow(),
        'updated_by': user.get('uid', '')
    }
    
    # Salvare istoric modificare
    modificare = {
        'tip': 'MODIFICARE',
        'suma_veche': existing.get('suma_curenta'),
        'suma_noua': Decimal128(suma_totala),
        'numar_document': data.get('numar_document'),
        'data_document': update_data['data_document'],
        'utilizator': user.get('uid', ''),
        'data_modificare': datetime.utcnow()
    }
    
    # Update
    result = db.alop_buget_aprobat.update_one(
        {'_id': ObjectId(item_id)},
        {
            '$set': update_data,
            '$push': {'modificari': modificare}
        }
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=500, detail="Eroare la actualizare")
    
    # Returnează item-ul actualizat
    updated = db.alop_buget_aprobat.find_one({'_id': ObjectId(item_id)})
    return format_buget_item(updated)

# ============================================================================
# BUGET APROBAT - SUPLIMENTARE / DIMINUARE
# ============================================================================

@router.post("/aprobat/{item_id}/ajustare")
async def ajustare_buget_aprobat(
    item_id: str,
    data: Dict[str, Any],
    user = Depends(verify_token)
) -> Dict[str, Any]:
    """
    Suplimentare sau diminuare buget
    Creează o nouă înregistrare de modificare
    """
    db = get_db()
    
    try:
        existing = db.alop_buget_aprobat.find_one({'_id': ObjectId(item_id)})
    except:
        raise HTTPException(status_code=400, detail="ID invalid")
    
    if not existing:
        raise HTTPException(status_code=404, detail="Înregistrare negăsită")
    
    # Validare
    tip_ajustare = data.get('tip_ajustare', '').upper()
    if tip_ajustare not in ['SUPLIMENTARE', 'DIMINUARE']:
        raise HTTPException(status_code=400, detail="Tipul ajustării trebuie să fie SUPLIMENTARE sau DIMINUARE")
    
    if not data.get('numar_document'):
        raise HTTPException(status_code=400, detail="Numărul documentului este obligatoriu")
    
    if not data.get('data_document'):
        raise HTTPException(status_code=400, detail="Data documentului este obligatorie")
    
    # Calculare sume
    t1 = Decimal(str(data.get('trimestru_1', 0) or 0))
    t2 = Decimal(str(data.get('trimestru_2', 0) or 0))
    t3 = Decimal(str(data.get('trimestru_3', 0) or 0))
    t4 = Decimal(str(data.get('trimestru_4', 0) or 0))
    
    suma_ajustare = t1 + t2 + t3 + t4
    if suma_ajustare == 0:
        suma_ajustare = Decimal(str(data.get('suma_totala', 0) or 0))
    
    if suma_ajustare <= 0:
        raise HTTPException(status_code=400, detail="Suma ajustării trebuie să fie pozitivă")
    
    # Calculare sumă nouă
    suma_curenta = Decimal(str(existing.get('suma_curenta', Decimal128('0')).to_decimal()))
    
    if tip_ajustare == 'SUPLIMENTARE':
        suma_noua = suma_curenta + suma_ajustare
        t1_nou = Decimal(str(existing.get('trimestre', {}).get('t1', Decimal128('0')).to_decimal())) + t1
        t2_nou = Decimal(str(existing.get('trimestre', {}).get('t2', Decimal128('0')).to_decimal())) + t2
        t3_nou = Decimal(str(existing.get('trimestre', {}).get('t3', Decimal128('0')).to_decimal())) + t3
        t4_nou = Decimal(str(existing.get('trimestre', {}).get('t4', Decimal128('0')).to_decimal())) + t4
    else:  # DIMINUARE
        if suma_ajustare > suma_curenta:
            raise HTTPException(status_code=400, detail="Suma diminuării nu poate depăși suma curentă")
        suma_noua = suma_curenta - suma_ajustare
        t1_nou = Decimal(str(existing.get('trimestre', {}).get('t1', Decimal128('0')).to_decimal())) - t1
        t2_nou = Decimal(str(existing.get('trimestre', {}).get('t2', Decimal128('0')).to_decimal())) - t2
        t3_nou = Decimal(str(existing.get('trimestre', {}).get('t3', Decimal128('0')).to_decimal())) - t3
        t4_nou = Decimal(str(existing.get('trimestre', {}).get('t4', Decimal128('0')).to_decimal())) - t4
    
    # Pregătire update
    data_doc = data.get('data_document')
    if isinstance(data_doc, str):
        data_doc = datetime.fromisoformat(data_doc)
    
    update_data = {
        'suma_curenta': Decimal128(suma_noua),
        'trimestre': {
            't1': Decimal128(t1_nou),
            't2': Decimal128(t2_nou),
            't3': Decimal128(t3_nou),
            't4': Decimal128(t4_nou)
        },
        'tip_operatie': tip_ajustare,
        'numar_document': data.get('numar_document'),
        'data_document': data_doc,
        'updated_at': datetime.utcnow(),
        'updated_by': user.get('uid', '')
    }
    
    # Salvare istoric
    modificare = {
        'tip': tip_ajustare,
        'suma_ajustare': Decimal128(suma_ajustare),
        'suma_veche': Decimal128(suma_curenta),
        'suma_noua': Decimal128(suma_noua),
        'trimestre_ajustare': {
            't1': Decimal128(t1),
            't2': Decimal128(t2),
            't3': Decimal128(t3),
            't4': Decimal128(t4)
        },
        'numar_document': data.get('numar_document'),
        'data_document': data_doc,
        'capitol_id': ObjectId(data['capitol_id']) if data.get('capitol_id') else existing.get('capitol_id'),
        'subcapitol_id': ObjectId(data['subcapitol_id']) if data.get('subcapitol_id') else existing.get('subcapitol_id'),
        'articol_id': ObjectId(data['articol_id']) if data.get('articol_id') else existing.get('articol_id'),
        'alineat_id': ObjectId(data['alineat_id']) if data.get('alineat_id') else existing.get('alineat_id'),
        'program_id': ObjectId(data['program_id']) if data.get('program_id') else None,
        'utilizator': user.get('uid', ''),
        'data_modificare': datetime.utcnow()
    }
    
    # Update
    result = db.alop_buget_aprobat.update_one(
        {'_id': ObjectId(item_id)},
        {
            '$set': update_data,
            '$push': {'modificari': modificare}
        }
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=500, detail="Eroare la actualizare")
    
    # Returnează item-ul actualizat
    updated = db.alop_buget_aprobat.find_one({'_id': ObjectId(item_id)})
    return format_buget_item(updated)

# ============================================================================
# EXPORT EXCEL
# ============================================================================

@router.get("/aprobat/export/excel")
async def export_buget_aprobat_excel(
    search: Optional[str] = None,
    capitol: Optional[str] = None,
    an_bugetar: Optional[int] = None,
    user = Depends(verify_token)
):
    """
    Exportă bugetul aprobat în format Excel
    """
    from fastapi.responses import StreamingResponse
    import io
    
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    except ImportError:
        raise HTTPException(status_code=500, detail="Modulul openpyxl nu este instalat")
    
    db = get_db()
    
    # Construire query
    query = {}
    if search:
        search_regex = re.compile(re.escape(search), re.IGNORECASE)
        query['$or'] = [
            {'cod_clasificare': search_regex},
            {'denumire': search_regex}
        ]
    if capitol:
        query['cod_clasificare_parsed.capitol'] = capitol
    if an_bugetar:
        query['an_bugetar'] = an_bugetar
    
    # Obținere date
    cursor = db.alop_buget_aprobat.find(query).sort('cod_clasificare', 1)
    
    # Creare workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Buget Aprobat"
    
    # Stiluri
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Header - conform documentației
    headers = [
        'Cod Clasificare',
        'Denumire',
        'Suma Inițială',
        'Suma Curentă',
        'Trimestrul 1',
        'Trimestrul 2',
        'Trimestrul 3',
        'Trimestrul 4',
        'Tip Operație',
        'Nr. Document',
        'Data Document',
        'Subunitate',
        'Program'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Date
    row = 2
    for item in cursor:
        ws.cell(row=row, column=1, value=item.get('cod_clasificare', '')).border = thin_border
        ws.cell(row=row, column=2, value=item.get('denumire', '')).border = thin_border
        ws.cell(row=row, column=3, value=decimal128_to_float(item.get('suma_initiala'))).border = thin_border
        ws.cell(row=row, column=4, value=decimal128_to_float(item.get('suma_curenta'))).border = thin_border
        ws.cell(row=row, column=5, value=decimal128_to_float(item.get('trimestre', {}).get('t1'))).border = thin_border
        ws.cell(row=row, column=6, value=decimal128_to_float(item.get('trimestre', {}).get('t2'))).border = thin_border
        ws.cell(row=row, column=7, value=decimal128_to_float(item.get('trimestre', {}).get('t3'))).border = thin_border
        ws.cell(row=row, column=8, value=decimal128_to_float(item.get('trimestre', {}).get('t4'))).border = thin_border
        ws.cell(row=row, column=9, value=item.get('tip_operatie', '')).border = thin_border
        ws.cell(row=row, column=10, value=item.get('numar_document', '')).border = thin_border
        
        data_doc = item.get('data_document')
        ws.cell(row=row, column=11, value=data_doc.strftime('%d.%m.%Y') if data_doc else '').border = thin_border
        
        ws.cell(row=row, column=12, value=item.get('cod_subunitate', '')).border = thin_border
        ws.cell(row=row, column=13, value=item.get('cod_program', '')).border = thin_border
        
        # Format numere
        for col in [3, 4, 5, 6, 7, 8]:
            ws.cell(row=row, column=col).number_format = '#,##0.00'
        
        row += 1
    
    # Ajustare lățime coloane
    column_widths = [18, 40, 15, 15, 15, 15, 15, 15, 15, 12, 12, 10, 10]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
    
    # Salvare în buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Returnare răspuns
    filename = f"buget_aprobat_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# ============================================================================
# NOMENCLATOARE - AJAX SEARCH
# ============================================================================

@router.get("/nomenclatoare/capitole")
async def search_capitole(
    q: Optional[str] = None,
    limit: int = Query(50, ge=1, le=200),
    user = Depends(verify_token)
) -> List[Dict[str, Any]]:
    """Căutare AJAX pentru capitole"""
    db = get_db()
    
    query = {}
    if q:
        search_regex = re.compile(re.escape(q), re.IGNORECASE)
        query['$or'] = [
            {'cod': search_regex},
            {'denumire': search_regex}
        ]
    
    cursor = db.alop_capitole.find(query).limit(limit)
    
    return [{
        '_id': str(item['_id']),
        'cod': item.get('cod', ''),
        'denumire': item.get('denumire', ''),
        'label': f"{item.get('cod', '')} - {item.get('denumire', '')}"
    } for item in cursor]

@router.get("/nomenclatoare/subcapitole")
async def search_subcapitole(
    q: Optional[str] = None,
    capitol: Optional[str] = None,
    limit: int = Query(50, ge=1, le=200),
    user = Depends(verify_token)
) -> List[Dict[str, Any]]:
    """Căutare AJAX pentru subcapitole"""
    db = get_db()
    
    query = {}
    if q:
        search_regex = re.compile(re.escape(q), re.IGNORECASE)
        query['$or'] = [
            {'cod': search_regex},
            {'denumire': search_regex}
        ]
    if capitol:
        query['cod'] = {'$regex': f'^{capitol}'}
    
    cursor = db.alop_subcapitole.find(query).limit(limit)
    
    return [{
        '_id': str(item['_id']),
        'cod': item.get('cod', ''),
        'denumire': item.get('denumire', ''),
        'label': f"{item.get('cod', '')} - {item.get('denumire', '')}"
    } for item in cursor]

@router.get("/nomenclatoare/aliniate")
async def search_aliniate(
    q: Optional[str] = None,
    limit: int = Query(50, ge=1, le=200),
    user = Depends(verify_token)
) -> List[Dict[str, Any]]:
    """Căutare AJAX pentru articole/aliniate"""
    db = get_db()
    
    query = {}
    if q:
        search_regex = re.compile(re.escape(q), re.IGNORECASE)
        query['$or'] = [
            {'cod': search_regex},
            {'denumire': search_regex}
        ]
    
    cursor = db.alop_aliniate.find(query).limit(limit)
    
    return [{
        '_id': str(item['_id']),
        'cod': item.get('cod', ''),
        'denumire': item.get('denumire', ''),
        'label': f"{item.get('cod', '')} - {item.get('denumire', '')}"
    } for item in cursor]

@router.get("/nomenclatoare/programe")
async def search_programe(
    q: Optional[str] = None,
    limit: int = Query(50, ge=1, le=200),
    user = Depends(verify_token)
) -> List[Dict[str, Any]]:
    """Căutare AJAX pentru programe"""
    db = get_db()
    
    query = {}
    if q:
        search_regex = re.compile(re.escape(q), re.IGNORECASE)
        query['$or'] = [
            {'cod': search_regex},
            {'denumire': search_regex}
        ]
    
    # Verifică dacă colecția există
    if 'alop_programe' not in db.list_collection_names():
        return []
    
    cursor = db.alop_programe.find(query).limit(limit)
    
    return [{
        '_id': str(item['_id']),
        'cod': item.get('cod', ''),
        'denumire': item.get('denumire', ''),
        'label': f"{item.get('cod', '')} - {item.get('denumire', '')}"
    } for item in cursor]
