"""
Routes pentru Angajamente Legale
Gestionează operațiunile CRUD pentru angajamentele legale
"""
from fastapi import APIRouter, HTTPException, Depends, Query, UploadFile, File
from fastapi.responses import FileResponse
from typing import Optional, Dict, Any, Tuple, List
from pydantic import BaseModel
from datetime import datetime, date
from decimal import Decimal
from bson import ObjectId, Decimal128
import re
import os
import unicodedata

try:
    from ..utils.db import get_db
    from ..routes.auth import verify_token
    from ..utils.logger import logger
    from ..utils.dataflows_docu import DataFlowsDocuClient
    from ..utils.file_handler import save_document_file, save_upload_file, get_file_path
except ImportError:
    from utils.db import get_db
    from routes.auth import verify_token
    from utils.logger import logger
    from utils.dataflows_docu import DataFlowsDocuClient
    from utils.file_handler import save_document_file, save_upload_file, get_file_path

router = APIRouter(prefix="/api/angajamente-legale", tags=["angajamente-legale"])

CODE_FIELD_CANDIDATES = ("cod", "cod_clasificare", "code")
NAME_FIELD_CANDIDATES = ("denumire", "nume", "name", "descriere")


def _normalize_text(value: Optional[str]) -> str:
    if not value:
        return ""
    normalized = unicodedata.normalize('NFD', str(value))
    stripped = ''.join(ch for ch in normalized if unicodedata.category(ch) != 'Mn')
    return stripped.lower().strip()


def _decimal_to_float(value: Any) -> float:
    try:
        return float(str(value))
    except Exception:
        return 0.0


def _format_date(value: Any) -> Optional[str]:
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d')
    if isinstance(value, date):
        return value.strftime('%Y-%m-%d')
    return value


class AngajamentGenerateRequest(BaseModel):
    template_code: Optional[str] = None
    template_name: Optional[str] = None


def _extract_first(doc: Dict[str, Any], keys: Tuple[str, ...]) -> Optional[str]:
    for key in keys:
        value = doc.get(key)
        if value:
            return str(value)
    return None


def _parse_next_number(last_value: Optional[str], default_width: int = 3) -> str:
    if not last_value:
        return str(1).zfill(default_width)
    match = re.match(r"\s*(\d+)", str(last_value))
    if not match:
        return str(1).zfill(default_width)
    num = int(match.group(1))
    width = max(len(match.group(1)), default_width)
    return str(num + 1).zfill(width)


def _normalize_object_id(value: Any) -> Optional[ObjectId]:
    if isinstance(value, ObjectId):
        return value
    if isinstance(value, str) and ObjectId.is_valid(value):
        return ObjectId(value)
    return None


def _parse_number_with_width(value: Any, default_width: int = 3) -> Tuple[Optional[int], int]:
    if value is None:
        return None, default_width
    match = re.match(r"\s*(\d+)", str(value))
    if not match:
        return None, default_width
    num = int(match.group(1))
    width = max(len(match.group(1)), default_width)
    return num, width


def _extract_config_start(db, slug: str) -> Tuple[Optional[int], int]:
    cfg = db.config.find_one({'slug': slug})
    if not cfg:
        return None, 3
    value = None
    for key in ("value", "content", "items", "number", "start", "start_value"):
        if key in cfg:
            value = cfg.get(key)
            break
    if isinstance(value, dict):
        for key in ("value", "number", "start", "start_value"):
            if key in value:
                value = value.get(key)
                break
    if isinstance(value, list) and value:
        value = value[0]
    num, width = _parse_number_with_width(value)
    return num, width

def parseaza_cod_clasificare(cod: str) -> dict:
    """Parsează codul de clasificare bugetară"""
    if not cod or len(cod) < 10:
        return {"capitol": "", "subcapitol": "", "articol": "", "alineat": ""}
    
    return {
        "capitol": cod[0:4],
        "subcapitol": cod[4:8],
        "articol": cod[8:10],
        "alineat": cod[10:] if len(cod) > 10 else ""
    }


@router.get("/next-number")
async def get_next_numar_angajament(
    an_bugetar: Optional[int] = None,
    current_user: dict = Depends(verify_token)
) -> Dict[str, Any]:
    """Return next available angajament legal number (without year)."""
    db = get_db(domain=current_user.get('domain'))
    try:
        query: Dict[str, Any] = {}
        if an_bugetar:
            query["an_bugetar"] = an_bugetar
        last = db.alop_angajamente_legale.find_one(query, sort=[("creat_la", -1)])
        last_num, last_width = _parse_number_with_width(last.get("numar_angajament") if last else None)
        start_num, start_width = _extract_config_start(db, "alop_angajamente_legale_start")
        width = max(last_width, start_width)
        if last_num is None:
            next_num = start_num if start_num is not None else 1
        else:
            if start_num is not None and last_num < start_num:
                next_num = start_num
            else:
                next_num = last_num + 1
        next_number = str(next_num).zfill(width)
        return {"next_number": next_number, "an_bugetar": an_bugetar}
    except Exception as e:
        logger.error(subject="Error computing next number", content=str(e), category="angajamente-legale")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/contracte")
async def get_contracte(
    search: Optional[str] = None,
    current_user: dict = Depends(verify_token)
) -> List[Dict[str, Any]]:
    """Get contracts for angajamente legale selection."""
    db = get_db(domain=current_user.get('domain'))
    try:
        query: Dict[str, Any] = {}
        if search:
            search_regex = {"$regex": search, "$options": "i"}
            query["$or"] = [
                {"numar_contract": search_regex},
                {"firma.denumire": search_regex},
                {"cod_firma": search_regex}
            ]
        contracte = list(db.alop_contracte.find(query).sort("data_contract", -1).limit(1000))
        result = []
        for contract in contracte:
            numar = contract.get("numar_contract", "")
            data = contract.get("data_contract")
            firma = contract.get("firma", {}).get("denumire") if isinstance(contract.get("firma"), dict) else ""
            label_parts = [numar, firma]
            label = " - ".join([p for p in label_parts if p])
            result.append({
                "value": str(contract["_id"]),
                "label": label or numar or str(contract["_id"]),
                "numar_contract": numar,
                "data_contract": data.isoformat() if isinstance(data, (datetime, date)) else data,
                "firma": firma
            })
        return result
    except Exception as e:
        logger.error(subject="Error fetching contracte", content=str(e), category="angajamente-legale")
        return []


@router.get("/referate/approved")
async def get_referate_approved(
    search: Optional[str] = None,
    current_user: dict = Depends(verify_token)
) -> List[Dict[str, Any]]:
    """Get list of approved referate for angajamente legale."""
    db = get_db(domain=current_user.get('domain'))
    try:
        query: Dict[str, Any] = {"stare": "Aprobat"}
        if search:
            search_regex = {"$regex": search, "$options": "i"}
            query["$or"] = [
                {"titlu": search_regex},
                {"nr": search_regex},
                {"departament": search_regex}
            ]
        referate = list(db.procurement_referate.find(query).sort("created_at", -1).limit(200))
        result = []
        for referat in referate:
            nr = referat.get("nr", "")
            titlu = referat.get("titlu", "")
            departament = referat.get("departament", "")
            label_parts = [f"#{nr}" if nr else "", titlu, departament]
            label = " - ".join([p for p in label_parts if p])
            result.append({
                "value": str(referat["_id"]),
                "label": label or titlu or str(referat["_id"]),
                "nr": nr,
                "titlu": titlu,
                "departament": departament,
                "valoare_estimata": referat.get("valoare_estimata"),
                "an_bugetar": referat.get("an_bugetar")
            })
        return result
    except Exception as e:
        logger.error(subject="Error fetching referate approved", content=str(e), category="angajamente-legale")
        return []


@router.get("/referate/{referat_id}")
async def get_referat_details(
    referat_id: str,
    current_user: dict = Depends(verify_token)
) -> Dict[str, Any]:
    """Get referat details for angajamente legale prefill."""
    db = get_db(domain=current_user.get('domain'))
    try:
        referat = db.procurement_referate.find_one({"_id": ObjectId(referat_id)})
        if not referat:
            raise HTTPException(status_code=404, detail="Referat not found")
        if referat.get("stare") != "Aprobat":
            raise HTTPException(status_code=400, detail="Referatul nu este aprobat")

        return {
            "id": str(referat.get("_id")),
            "nr": referat.get("nr"),
            "titlu": referat.get("titlu"),
            "departament": referat.get("departament"),
            "categorie": referat.get("categorie"),
            "justificare": referat.get("justificare"),
            "valoare_estimata": referat.get("valoare_estimata"),
            "surse_finantare": referat.get("surse_finantare"),
            "fonduri_disponibile": referat.get("fonduri_disponibile"),
            "an_bugetar": referat.get("an_bugetar"),
            "bunuri_servicii": referat.get("bunuri_servicii", []),
            "articol": referat.get("articol") or referat.get("articol_bugetar")
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(subject="Error fetching referat details", content=str(e), category="angajamente-legale")
        raise HTTPException(status_code=500, detail=str(e))


def actualizeaza_cont_angajament_legal(
    cod_clasificare: str,
    cod_subunitate: str,
    cod_beneficiar: str,
    suma: Decimal,
    tip: str,  # "ADAUGARE" sau "DIMINUARE"
    db=None
):
    """
    Actualizează contul 80670 pentru angajamente legale.
    
    În sistemul ALOP, contul 80670 ține evidența angajamentelor legale.
    """
    if db is None:
        db = get_db()
    cont = "80670"
    
    if tip == "ADAUGARE":
        increment = Decimal128(suma)
    else:
        increment = Decimal128(-suma)
    
    try:
        db.alop_conturi_sume.update_one(
            {
                "cont": cont,
                "cod_clasificare": cod_clasificare,
                "cod_subunitate": cod_subunitate,
                "repartitor": cod_beneficiar
            },
            {
                "$inc": {"sold": increment},
                "$set": {
                    "actualizat_la": datetime.utcnow(),
                    "cont_denumire": "Angajamente legale"
                },
                "$setOnInsert": {
                    "creat_la": datetime.utcnow()
                }
            },
            upsert=True
        )
        
        logger.info(
            subject=f"Cont 80670 actualizat",
            content=f"Clasificare: {cod_clasificare}, Suma: {suma}, Tip: {tip}",
            category="conturi"
        )
    except Exception as e:
        logger.error(
            subject=f"Eroare actualizare cont 80670",
            content=str(e),
            category="conturi"
        )


@router.get("")
async def get_angajamente_legale(
    page: int = Query(1, ge=1),
    limit: int = Query(50, ge=1, le=500),
    search: Optional[str] = None,
    an_bugetar: Optional[int] = None,
    stare: Optional[str] = None,
    cod_beneficiar: Optional[str] = None,
    angajament_bugetar_id: Optional[str] = None,
    sort_by: str = "data_angajament",
    sort_order: str = "desc",
    user = Depends(verify_token)
):
    """Obține lista angajamentelor legale cu paginare și filtre"""
    db = get_db(domain=user.get('domain'))
    
    try:
        # Construire query
        query_filter = {}
        
        if search:
            search_regex = {"$regex": search, "$options": "i"}
            query_filter["$or"] = [
                {"numar_angajament": search_regex},
                {"scop": search_regex},
                {"beneficiar.denumire": search_regex},
                {"contract.numar": search_regex}
            ]
        
        if an_bugetar:
            query_filter["an_bugetar"] = an_bugetar
        
        if stare:
            query_filter["stare"] = stare
        
        if cod_beneficiar:
            query_filter["cod_beneficiar"] = cod_beneficiar
        
        if angajament_bugetar_id:
            query_filter["angajament_bugetar_id"] = ObjectId(angajament_bugetar_id)
        
        # Count total
        total = db.alop_angajamente_legale.count_documents(query_filter)
        
        # Calculate skip
        skip = (page - 1) * limit
        
        # Sort direction
        sort_direction = -1 if sort_order == "desc" else 1
        
        # Query
        cursor = db.alop_angajamente_legale.find(query_filter).sort(sort_by, sort_direction).skip(skip).limit(limit)
        
        items = []
        for item in cursor:
            item["_id"] = str(item["_id"])
            if item.get("angajament_bugetar_id"):
                item["angajament_bugetar_id"] = str(item["angajament_bugetar_id"])
            
            # Convert Decimal128 to float
            for key in ["suma_lei", "suma_valuta", "suma_ordonantata", "suma_disponibila"]:
                if key in item and item[key]:
                    item[key] = float(str(item[key]))
            items.append(item)
        
        return {
            "items": items,
            "total": total,
            "page": page,
            "limit": limit,
            "pages": (total + limit - 1) // limit
        }
    
    except Exception as e:
        logger.error(subject="Error fetching angajamente legale", content=str(e), category="angajamente-legale")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/{angajament_id}")
async def get_angajament_legal(
    angajament_id: str,
    user = Depends(verify_token)
):
    """Obține detaliile unui angajament legal"""
    db = get_db(domain=user.get('domain'))
    
    try:
        angajament = db.alop_angajamente_legale.find_one({"_id": ObjectId(angajament_id)})
        
        if not angajament:
            raise HTTPException(status_code=404, detail="Angajamentul legal nu există")
        
        angajament["_id"] = str(angajament["_id"])
        if angajament.get("angajament_bugetar_id"):
            angajament["angajament_bugetar_id"] = str(angajament["angajament_bugetar_id"])
        
        # Convert Decimal128 to float
        for key in ["suma_lei", "suma_valuta", "suma_ordonantata", "suma_disponibila"]:
            if key in angajament and angajament[key]:
                angajament[key] = float(str(angajament[key]))
        
        return angajament
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(subject=f"Error fetching angajament legal {angajament_id}", content=str(e), category="angajamente-legale")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("")
async def create_angajament_legal(
    data: Dict[str, Any],
    user = Depends(verify_token)
):
    """
    Creează un nou angajament legal
    
    Validări:
    - Suma nu poate depăși disponibilul din angajamentul bugetar
    - Beneficiarul trebuie să coincidă cu cel din angajamentul bugetar
    """
    db = get_db(domain=user.get('domain'))
    
    try:
        # Extragere date
        angajament_bugetar_id = data.get("angajament_bugetar_id")
        contract_id = data.get("contract_id")
        suma_lei = Decimal(str(data.get("suma_lei", 0)))
        numar_angajament = data.get("numar_angajament")
        data_angajament = data.get("data_angajament")
        scop = data.get("scop")
        referat_id = data.get("referat_id")
        referat_info = None

        # Prefill from referat (if provided)
        referat_doc = None
        if referat_id and ObjectId.is_valid(referat_id):
            referat_doc = db.procurement_referate.find_one({"_id": ObjectId(referat_id)})
            if not referat_doc:
                raise HTTPException(status_code=404, detail="Referatul nu exist??")
            if referat_doc.get("stare") != "Aprobat":
                raise HTTPException(status_code=400, detail="Referatul nu este aprobat")

            if (suma_lei <= 0) and referat_doc.get("valoare_estimata"):
                suma_lei = Decimal(str(referat_doc.get("valoare_estimata")))
            if not scop:
                scop = referat_doc.get("titlu") or referat_doc.get("justificare") or scop

            referat_info = {
                "referat_id": str(referat_doc.get("_id")),
                "nr": referat_doc.get("nr"),
                "titlu": referat_doc.get("titlu"),
                "departament": referat_doc.get("departament"),
                "categorie": referat_doc.get("categorie"),
                "valoare_estimata": referat_doc.get("valoare_estimata"),
                "an_bugetar": referat_doc.get("an_bugetar"),
                "articol": referat_doc.get("articol") or referat_doc.get("articol_bugetar")
            }
        
        # Valid?ri
        if not all([angajament_bugetar_id, suma_lei > 0, numar_angajament]):
            raise HTTPException(status_code=400, detail="C?mpuri obligatorii lips??")

# Verificare angajament bugetar
        angajament_bugetar = db.alop_angajamente_bugetare.find_one({"_id": ObjectId(angajament_bugetar_id)})
        
        if not angajament_bugetar:
            raise HTTPException(status_code=404, detail="Angajamentul bugetar nu există")
        
        if angajament_bugetar["anulat"]:
            raise HTTPException(status_code=400, detail="Angajamentul bugetar este anulat")
        
        # Verificare disponibil
        disponibil_abu = Decimal(str(angajament_bugetar["suma_disponibila"]))
        
        if suma_lei > disponibil_abu:
            raise HTTPException(
                status_code=400,
                detail=f"Suma ({suma_lei}) depășește disponibilul din angajamentul bugetar ({disponibil_abu})"
            )
        
        # Extragere date din angajament bugetar
        cod_clasificare = angajament_bugetar.get("cod_clasificare")
        cod_subunitate = angajament_bugetar.get("cod_subunitate")
        clasificare_id = angajament_bugetar.get("clasificare_id")
        subunitate_id = angajament_bugetar.get("subunitate_id")
        if not cod_clasificare and clasificare_id:
            try:
                doc = db.alop_clasificari.find_one({"_id": _normalize_object_id(clasificare_id)})
                cod_clasificare = _extract_first(doc or {}, CODE_FIELD_CANDIDATES)
            except Exception:
                cod_clasificare = None
        if not cod_subunitate and subunitate_id:
            try:
                doc = db.alop_subunitati.find_one({"_id": _normalize_object_id(subunitate_id)})
                cod_subunitate = _extract_first(doc or {}, CODE_FIELD_CANDIDATES)
            except Exception:
                cod_subunitate = None

        if not cod_clasificare or not cod_subunitate:
            raise HTTPException(status_code=400, detail="Clasificarea sau subunitatea lipsesc pentru angajamentul bugetar")
        an_bugetar = angajament_bugetar["an_bugetar"]
        cod_beneficiar = angajament_bugetar["cod_beneficiar"]
        
        # Lookup beneficiar
        beneficiar = db.alop_firme.find_one({"cod": cod_beneficiar})
        beneficiar_data = None
        if beneficiar:
            beneficiar_data = {
                "denumire": beneficiar.get("denumire", ""),
                "cif": beneficiar.get("cif", ""),
                "adresa": beneficiar.get("adresa_completa") or beneficiar.get("adresa", ""),
                "iban": beneficiar.get("iban", "")
            }
        
        # Verificare unicitate
        existing = db.alop_angajamente_legale.find_one({
            "numar_angajament": numar_angajament,
            "an_bugetar": an_bugetar
        })
        
        if existing:
            raise HTTPException(status_code=400, detail=f"Angajamentul legal {numar_angajament} există deja")
        
        # Parsare data
        if isinstance(data_angajament, str):
            data_angajament = datetime.fromisoformat(data_angajament.replace('Z', '+00:00'))
        elif isinstance(data_angajament, date):
            data_angajament = datetime.combine(data_angajament, datetime.min.time())
        
        # Contract info
        contract_info = None
        if contract_id and ObjectId.is_valid(contract_id):
            contract = db.alop_contracte.find_one({"_id": ObjectId(contract_id)})
            if contract:
                contract_info = {
                    "numar": contract.get("numar_contract"),
                    "data": contract.get("data_contract"),
                    "explicatie": contract.get("explicatie", ""),
                    "contract_id": contract_id
                }
        if not contract_info and data.get("numar_contract"):
            contract_info = {
                "numar": data.get("numar_contract"),
                "data": datetime.fromisoformat(data.get("data_contract").replace('Z', '+00:00')) if data.get("data_contract") else None,
                "explicatie": data.get("contract_explicatie", "")
            }
        
        # Construire document
        document = {
            "numar_angajament": numar_angajament,
            "data_angajament": data_angajament,
            "an_bugetar": an_bugetar,
            
            "angajament_bugetar_id": ObjectId(angajament_bugetar_id),
            "angajament_bugetar_numar": angajament_bugetar["numar_angajament"],
            
            "cod_clasificare": cod_clasificare,
            "cod_clasificare_parsed": parseaza_cod_clasificare(cod_clasificare),
            
            "suma_lei": Decimal128(suma_lei),
            "suma_valuta": Decimal128(Decimal(str(data.get("suma_valuta", 0)))),
            "cod_moneda": data.get("cod_moneda", "LEI"),
            "suma_ordonantata": Decimal128(Decimal(0)),
            "suma_disponibila": Decimal128(suma_lei),
            
            "cod_beneficiar": cod_beneficiar,
            "beneficiar": beneficiar_data,
            
            "contract": contract_info,
            "referat_id": referat_info.get("referat_id") if referat_info else None,
            "referat_info": referat_info,
            
            "scop": scop,
            "cont_contabil": data.get("cont_contabil"),
            "cod_compartiment": data.get("cod_compartiment"),
            "clasificare_id": _normalize_object_id(clasificare_id) if clasificare_id else None,
            "subunitate_id": _normalize_object_id(subunitate_id) if subunitate_id else None,
            "cod_subunitate": cod_subunitate,
            "cod_program": data.get("cod_program"),
            
            "stare": "ACTIV",
            "tip_operatie": "ANGAJARE",
            "anulat": False,
            "motiv_anulare": None,
            "data_anulare": None,

            "generated_docs": [],
            "signed_pdf_hash": None,
            "signed_pdf_filename": None,
            "signed_pdf_uploaded_at": None,
            "signed_pdf_uploaded_by": None,
            
            "ordonantari": [],
            
            "creat_la": datetime.utcnow(),
            "actualizat_la": datetime.utcnow(),
            "utilizator_creare": user.get("username", ""),
            "utilizator_actualizare": user.get("username", "")
        }
        
        result = db.alop_angajamente_legale.insert_one(document)
        
        # Actualizare angajament bugetar
        db.alop_angajamente_bugetare.update_one(
            {"_id": ObjectId(angajament_bugetar_id)},
            {
                "$inc": {
                    "suma_consumata": Decimal128(suma_lei),
                    "suma_disponibila": Decimal128(-suma_lei)
                },
                "$push": {
                    "angajamente_legale": {
                        "angajament_legal_id": result.inserted_id,
                        "numar": numar_angajament,
                        "suma": Decimal128(suma_lei),
                        "data": data_angajament
                    }
                }
            }
        )
        
        logger.info(
            subject="Angajament legal creat",
            content=f"Nr: {numar_angajament}, Suma: {suma_lei}, User: {user.get('username')}",
            category="angajamente-legale"
        )
        
        return {
            "success": True,
            "id": str(result.inserted_id),
            "numar_angajament": numar_angajament
        }
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(subject="Error creating angajament legal", content=str(e), category="angajamente-legale")
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/{angajament_id}")
async def delete_angajament_legal(
    angajament_id: str,
    motiv: str = Query(..., description="Motivul anulării"),
    user = Depends(verify_token)
):
    """Anulează un angajament legal"""
    db = get_db(domain=user.get('domain'))
    
    try:
        angajament = db.alop_angajamente_legale.find_one({"_id": ObjectId(angajament_id)})
        
        if not angajament:
            raise HTTPException(status_code=404, detail="Angajamentul legal nu există")
        
        if angajament["anulat"]:
            raise HTTPException(status_code=400, detail="Angajamentul este deja anulat")
        
        # Verificare ordonanțări
        suma_ordonantata = Decimal(str(angajament["suma_ordonantata"]))
        if suma_ordonantata > 0:
            raise HTTPException(
                status_code=400,
                detail=f"Nu se poate anula - există ordonanțări în valoare de {suma_ordonantata}"
            )
        
        suma_angajament = Decimal(str(angajament["suma_lei"]))
        
        # Anulare
        result = db.alop_angajamente_legale.update_one(
            {"_id": ObjectId(angajament_id)},
            {
                "$set": {
                    "anulat": True,
                    "stare": "ANULAT",
                    "motiv_anulare": motiv,
                    "data_anulare": datetime.utcnow(),
                    "actualizat_la": datetime.utcnow(),
                    "utilizator_actualizare": user.get("username", "")
                }
            }
        )
        
        # Eliberare sumă în angajamentul bugetar
        if angajament.get("angajament_bugetar_id"):
            db.alop_angajamente_bugetare.update_one(
                {"_id": angajament["angajament_bugetar_id"]},
                {
                    "$inc": {
                        "suma_consumata": Decimal128(-suma_angajament),
                        "suma_disponibila": Decimal128(suma_angajament)
                    },
                    "$pull": {
                        "angajamente_legale": {
                            "angajament_legal_id": ObjectId(angajament_id)
                        }
                    }
                }
            )
        
        logger.info(
            subject="Angajament legal anulat",
            content=f"ID: {angajament_id}, Motiv: {motiv}, User: {user.get('username')}",
            category="angajamente-legale"
        )
        
        return {"success": True, "modified_count": result.modified_count}
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(subject=f"Error deleting angajament legal {angajament_id}", content=str(e), category="angajamente-legale")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/export/excel")
async def export_angajamente_legale_excel(
    an_bugetar: Optional[int] = None,
    stare: Optional[str] = None,
    search: Optional[str] = None,
    user = Depends(verify_token)
):
    """Export angajamente legale în format Excel"""
    from fastapi.responses import StreamingResponse
    import io
    
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")
    
    db = get_db(domain=user.get('domain'))
    
    try:
        # Construire query
        query_filter = {}
        
        if search:
            search_regex = {"$regex": search, "$options": "i"}
            query_filter["$or"] = [
                {"numar_angajament": search_regex},
                {"scop": search_regex},
                {"beneficiar.denumire": search_regex}
            ]
        
        if an_bugetar:
            query_filter["an_bugetar"] = an_bugetar
        
        if stare:
            query_filter["stare"] = stare
        
        # Query
        cursor = db.alop_angajamente_legale.find(query_filter).sort("data_angajament", -1).limit(10000)
        
        # Creare workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Angajamente Legale"
        
        # Header styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Headers
        headers = [
            "Nr. Angajament",
            "Data",
            "An Bugetar",
            "Cod Clasificare",
            "Beneficiar",
            "CIF",
            "Sumă Lei",
            "Sumă Ordonanțată",
            "Disponibil",
            "Stare",
            "Contract",
            "Scop",
            "Ang. Bugetar"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Data rows
        row = 2
        for item in cursor:
            ws.cell(row=row, column=1, value=item.get("numar_angajament", ""))
            ws.cell(row=row, column=2, value=item.get("data_angajament").strftime("%d.%m.%Y") if item.get("data_angajament") else "")
            ws.cell(row=row, column=3, value=item.get("an_bugetar", ""))
            ws.cell(row=row, column=4, value=item.get("cod_clasificare", ""))
            
            beneficiar = item.get("beneficiar", {})
            ws.cell(row=row, column=5, value=beneficiar.get("denumire", "") if beneficiar else "")
            ws.cell(row=row, column=6, value=beneficiar.get("cif", "") if beneficiar else "")
            
            ws.cell(row=row, column=7, value=float(str(item.get("suma_lei", 0))))
            ws.cell(row=row, column=8, value=float(str(item.get("suma_ordonantata", 0))))
            ws.cell(row=row, column=9, value=float(str(item.get("suma_disponibila", 0))))
            ws.cell(row=row, column=10, value=item.get("stare", ""))
            
            contract = item.get("contract", {})
            ws.cell(row=row, column=11, value=contract.get("numar", "") if contract else "")
            ws.cell(row=row, column=12, value=item.get("scop", ""))
            ws.cell(row=row, column=13, value=item.get("angajament_bugetar_numar", ""))
            
            row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Return as streaming response
        filename = f"angajamente_legale_{an_bugetar or 'toate'}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    
    except Exception as e:
        logger.error(subject="Error exporting angajamente legale to Excel", content=str(e), category="angajamente-legale")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/{angajament_id}/documents/generate")
async def generate_angajament_legal_document(
    angajament_id: str,
    request: AngajamentGenerateRequest,
    current_user: dict = Depends(verify_token)
):
    """Generate angajament legal document using DataFlows Docu."""
    db = get_db(domain=current_user.get('domain'))

    try:
        angajament = db.alop_angajamente_legale.find_one({"_id": ObjectId(angajament_id)})
        if not angajament:
            raise HTTPException(status_code=404, detail="Angajamentul nu există")

        template_code = request.template_code or "P3C6R2ENXSIC"
        template_name = request.template_name or "Angajament Legal"

        beneficiar = angajament.get("beneficiar") or {}
        contract_info = angajament.get("contract") or {}
        cod_subunitate = angajament.get("cod_subunitate")
        if not cod_subunitate and angajament.get("subunitate_id"):
            try:
                doc = db.alop_subunitati.find_one({"_id": ObjectId(str(angajament.get("subunitate_id")))})
                cod_subunitate = _extract_first(doc or {}, CODE_FIELD_CANDIDATES)
            except Exception:
                cod_subunitate = None
        payload_data = {
            "data": {
                "id": str(angajament.get("_id")),
                "numar_angajament": angajament.get("numar_angajament"),
                "data_angajament": _format_date(angajament.get("data_angajament")),
                "an_bugetar": angajament.get("an_bugetar"),
                "angajament_bugetar_numar": angajament.get("angajament_bugetar_numar"),
                "cod_clasificare": angajament.get("cod_clasificare"),
                "cod_subunitate": cod_subunitate,
                "cod_beneficiar": angajament.get("cod_beneficiar"),
                "beneficiar": beneficiar,
                "suma_lei": _decimal_to_float(angajament.get("suma_lei")),
                "suma_valuta": _decimal_to_float(angajament.get("suma_valuta")),
                "suma_disponibila": _decimal_to_float(angajament.get("suma_disponibila")),
                "cod_moneda": angajament.get("cod_moneda", "LEI"),
                "scop": angajament.get("scop"),
                "contract": {
                    "numar": contract_info.get("numar"),
                    "data": _format_date(contract_info.get("data")),
                    "explicatie": contract_info.get("explicatie")
                },
                "referat": angajament.get("referat_info") or {}
            },
            "generated_at": datetime.utcnow().isoformat(),
            "generated_by": current_user.get("username", "unknown")
        }

        client = DataFlowsDocuClient()
        filename = f"angajament_legal_{angajament.get('numar_angajament', '')}_{angajament_id[:8]}"
        document_bytes = client.create_realtime_job(
            template_code=template_code,
            data=payload_data,
            format="pdf",
            filename=filename
        )

        if not document_bytes:
            raise HTTPException(status_code=500, detail="Failed to generate document")

        file_hash = save_document_file(document_bytes, f"{filename}.pdf")
        generated_at = datetime.utcnow()

        generated_docs = list(angajament.get("generated_docs", []))
        doc_id = str(ObjectId())
        updated_entry = {
            "id": doc_id,
            "template_code": template_code,
            "template_name": template_name,
            "file_hash": file_hash,
            "filename": f"{filename}.pdf",
            "generated_at": generated_at,
            "generated_by": current_user.get("username", "unknown")
        }

        replaced = False
        for idx, entry in enumerate(generated_docs):
            if entry.get("template_code") == template_code:
                updated_entry["id"] = entry.get("id", doc_id)
                generated_docs[idx] = updated_entry
                replaced = True
                break

        if not replaced:
            generated_docs.append(updated_entry)

        db.alop_angajamente_legale.update_one(
            {"_id": ObjectId(angajament_id)},
            {"$set": {"generated_docs": generated_docs, "actualizat_la": datetime.utcnow()}}
        )

        updated_entry["generated_at"] = generated_at.isoformat()

        return {"success": True, "document": updated_entry}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(subject="Error generating angajament legal document", content=str(e), category="angajamente-legale")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/{angajament_id}/documents/{doc_entry_id}/download")
async def download_angajament_legal_document(
    angajament_id: str,
    doc_entry_id: str,
    current_user: dict = Depends(verify_token)
):
    """Download generated angajament legal document."""
    db = get_db(domain=current_user.get('domain'))

    try:
        angajament = db.alop_angajamente_legale.find_one({"_id": ObjectId(angajament_id)})
        if not angajament:
            raise HTTPException(status_code=404, detail="Angajamentul nu există")

        docs = angajament.get("generated_docs", [])
        entry = next((doc for doc in docs if doc.get("id") == doc_entry_id), None)
        if not entry or not entry.get("file_hash"):
            raise HTTPException(status_code=404, detail="Document not found")

        file_path = get_file_path(entry.get("file_hash"))
        if not file_path or not os.path.exists(file_path):
            raise HTTPException(status_code=404, detail="File not found")

        filename = entry.get("filename", f"angajament_legal_{angajament_id}.pdf")
        return FileResponse(
            file_path,
            media_type="application/pdf",
            headers={"Content-Disposition": f'inline; filename=\"{filename}\"'}
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(subject="Error downloading angajament legal document", content=str(e), category="angajamente-legale")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/{angajament_id}/signed/upload")
async def upload_signed_angajament_legal(
    angajament_id: str,
    file: UploadFile = File(...),
    current_user: dict = Depends(verify_token)
):
    """Upload signed angajament legal PDF."""
    db = get_db(domain=current_user.get('domain'))

    try:
        angajament = db.alop_angajamente_legale.find_one({"_id": ObjectId(angajament_id)})
        if not angajament:
            raise HTTPException(status_code=404, detail="Angajamentul nu există")

        if not file.filename or not file.filename.lower().endswith(".pdf"):
            raise HTTPException(status_code=400, detail="Doar fișiere PDF sunt permise")

        metadata = await save_upload_file(file)
        now = datetime.utcnow()

        db.alop_angajamente_legale.update_one(
            {"_id": ObjectId(angajament_id)},
            {"$set": {
                "signed_pdf_hash": metadata.get("hash"),
                "signed_pdf_filename": metadata.get("original_filename", file.filename),
                "signed_pdf_uploaded_at": now,
                "signed_pdf_uploaded_by": current_user.get("username", "unknown"),
                "actualizat_la": now
            }}
        )

        return {
            "success": True,
            "file_hash": metadata.get("hash"),
            "filename": metadata.get("original_filename", file.filename)
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(subject="Error uploading signed angajament legal", content=str(e), category="angajamente-legale")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/{angajament_id}/signed/download")
async def download_signed_angajament_legal(
    angajament_id: str,
    current_user: dict = Depends(verify_token)
):
    """Download signed angajament legal PDF."""
    db = get_db(domain=current_user.get('domain'))

    try:
        angajament = db.alop_angajamente_legale.find_one({"_id": ObjectId(angajament_id)})
        if not angajament:
            raise HTTPException(status_code=404, detail="Angajamentul nu există")

        file_hash = angajament.get("signed_pdf_hash")
        if not file_hash:
            raise HTTPException(status_code=404, detail="Signed document not found")

        file_path = get_file_path(file_hash)
        if not file_path or not os.path.exists(file_path):
            raise HTTPException(status_code=404, detail="File not found")

        filename = angajament.get("signed_pdf_filename", f"angajament_legal_{angajament_id}_signed.pdf")
        return FileResponse(
            file_path,
            media_type="application/pdf",
            headers={"Content-Disposition": f'inline; filename=\"{filename}\"'}
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(subject="Error downloading signed angajament legal", content=str(e), category="angajamente-legale")
        raise HTTPException(status_code=500, detail=str(e))
