"""
Routes pentru Angajamente Bugetare
Gestionează operațiunile CRUD pentru angajamentele bugetare
"""
from fastapi import APIRouter, HTTPException, Depends, Query, UploadFile, File
from fastapi.responses import FileResponse
from typing import Optional, List, Dict, Any, Tuple
from pydantic import BaseModel
from datetime import datetime, date
from decimal import Decimal
from bson import ObjectId, Decimal128
import re
import os
import unicodedata

try:
    from ..utils.db import get_db
    from ..routes.auth import verify_token
    from ..utils.logger import logger
    from ..utils.dataflows_docu import DataFlowsDocuClient
    from ..utils.file_handler import save_document_file, save_upload_file, get_file_path
    from ..utils.permission_helpers import is_admin_user
except ImportError:
    from utils.db import get_db
    from routes.auth import verify_token
    from utils.logger import logger
    from utils.dataflows_docu import DataFlowsDocuClient
    from utils.file_handler import save_document_file, save_upload_file, get_file_path
    from utils.permission_helpers import is_admin_user

router = APIRouter(prefix="/api/angajamente-bugetare", tags=["angajamente-bugetare"])

CODE_FIELD_CANDIDATES = ("cod", "cod_clasificare", "code")
NAME_FIELD_CANDIDATES = ("denumire", "nume", "name", "descriere")


def _normalize_text(value: Optional[str]) -> str:
    if not value:
        return ""
    normalized = unicodedata.normalize('NFD', str(value))
    stripped = ''.join(ch for ch in normalized if unicodedata.category(ch) != 'Mn')
    return stripped.lower().strip()


def _is_economic_department(user: dict) -> bool:
    departament = _normalize_text(user.get('departament') or user.get('department'))
    if not departament:
        return False
    return departament in {
        'contabilitate',
        'buget',
        'finante',
        'buget finante',
        'buget/finante',
        'economic',
        'directia economica'
    }


def _decimal_to_float(value: Any) -> float:
    try:
        return float(str(value))
    except Exception:
        return 0.0


def _format_date(value: Any) -> Optional[str]:
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d')
    if isinstance(value, date):
        return value.strftime('%Y-%m-%d')
    return value


class AngajamentGenerateRequest(BaseModel):
    template_code: Optional[str] = None
    template_name: Optional[str] = None


def _extract_first(doc: Dict[str, Any], keys: Tuple[str, ...]) -> Optional[str]:
    for key in keys:
        value = doc.get(key)
        if value:
            return str(value)
    return None


def _resolve_reference(
    db,
    collection: str,
    ref_id: Optional[str],
    fallback_code: Optional[str],
    code_keys: Tuple[str, ...]
) -> Tuple[Optional[ObjectId], Optional[str], Optional[str]]:
    """Return (oid, code, label_part) using id or code fallback."""
    if ref_id:
        try:
            doc = db[collection].find_one({"_id": ObjectId(ref_id)})
            if doc:
                code = _extract_first(doc, code_keys)
                name = _extract_first(doc, NAME_FIELD_CANDIDATES)
                return ObjectId(ref_id), code, name
        except Exception:
            pass
    if fallback_code and ObjectId.is_valid(fallback_code):
        try:
            doc = db[collection].find_one({"_id": ObjectId(fallback_code)})
            if doc:
                code = _extract_first(doc, code_keys)
                name = _extract_first(doc, NAME_FIELD_CANDIDATES)
                return ObjectId(fallback_code), code, name
        except Exception:
            pass
    if fallback_code:
        for key in code_keys:
            try:
                doc = db[collection].find_one({key: fallback_code})
                if doc:
                    code = _extract_first(doc, code_keys)
                    name = _extract_first(doc, NAME_FIELD_CANDIDATES)
                    return doc.get("_id"), code, name
            except Exception:
                continue
    if fallback_code:
        return None, fallback_code, None
    return None, None, None


def _parse_next_number(last_value: Optional[str], default_width: int = 3) -> str:
    if not last_value:
        return str(1).zfill(default_width)
    match = re.match(r"\s*(\d+)", str(last_value))
    if not match:
        return str(1).zfill(default_width)
    num = int(match.group(1))
    width = max(len(match.group(1)), default_width)
    return str(num + 1).zfill(width)


def _normalize_object_id(value: Any) -> Optional[ObjectId]:
    if isinstance(value, ObjectId):
        return value
    if isinstance(value, str) and ObjectId.is_valid(value):
        return ObjectId(value)
    return None


def _parse_number_with_width(value: Any, default_width: int = 3) -> Tuple[Optional[int], int]:
    if value is None:
        return None, default_width
    match = re.match(r"\s*(\d+)", str(value))
    if not match:
        return None, default_width
    num = int(match.group(1))
    width = max(len(match.group(1)), default_width)
    return num, width


def _extract_config_start(db, slug: str) -> Tuple[Optional[int], int]:
    cfg = db.config.find_one({'slug': slug})
    if not cfg:
        return None, 3
    value = None
    for key in ("value", "content", "items", "number", "start", "start_value"):
        if key in cfg:
            value = cfg.get(key)
            break
    if isinstance(value, dict):
        for key in ("value", "number", "start", "start_value"):
            if key in value:
                value = value.get(key)
                break
    if isinstance(value, list) and value:
        value = value[0]
    num, width = _parse_number_with_width(value)
    return num, width

def parseaza_cod_clasificare(cod: str) -> dict:
    """Parsează codul de clasificare bugetară"""
    if not cod or len(cod) < 10:
        return {"capitol": "", "subcapitol": "", "articol": "", "alineat": ""}
    
    return {
        "capitol": cod[0:4],
        "subcapitol": cod[4:8],
        "articol": cod[8:10],
        "alineat": cod[10:] if len(cod) > 10 else ""
    }


def actualizeaza_cont_angajament_bugetar(
    cod_clasificare: str,
    cod_subunitate: str,
    cod_beneficiar: str,
    suma: Decimal,
    tip: str,  # "ADAUGARE" sau "DIMINUARE"
    db=None
):
    """
    Actualizează contul 80660 pentru angajamente bugetare.
    
    În sistemul ALOP, contul 80660 ține evidența angajamentelor bugetare.
    """
    if db is None:
        db = get_db()
    cont = "80660"
    
    if tip == "ADAUGARE":
        increment = Decimal128(suma)
    else:
        increment = Decimal128(-suma)
    
    try:
        db.alop_conturi_sume.update_one(
            {
                "cont": cont,
                "cod_clasificare": cod_clasificare,
                "cod_subunitate": cod_subunitate,
                "repartitor": cod_beneficiar
            },
            {
                "$inc": {"sold": increment},
                "$set": {
                    "actualizat_la": datetime.utcnow(),
                    "cont_denumire": "Angajamente bugetare"
                },
                "$setOnInsert": {
                    "creat_la": datetime.utcnow()
                }
            },
            upsert=True
        )
        
        logger.info(
            subject=f"Cont 80660 actualizat",
            content=f"Clasificare: {cod_clasificare}, Suma: {suma}, Tip: {tip}",
            category="conturi"
        )
    except Exception as e:
        logger.error(
            subject=f"Eroare actualizare cont 80660",
            content=str(e),
            category="conturi"
        )


async def calculeaza_disponibil_buget(
    cod_clasificare: str,
    cod_subunitate: str,
    an_bugetar: int,
    exclude_angajament_id: Optional[str] = None,
    db=None
) -> Decimal:
    """Calculează disponibilul din buget pentru o clasificare"""
    if db is None:
        db = get_db()
    
    # Buget aprobat
    buget = db.alop_buget_aprobat.find_one({
        "cod_clasificare": cod_clasificare,
        "cod_subunitate": cod_subunitate,
        "an_bugetar": an_bugetar
    })
    
    if not buget:
        return Decimal(0)
    
    suma_aprobata = Decimal(str(buget["suma_curenta"]))
    
    # Angajamente bugetare existente
    match_filter = {
        "cod_clasificare": cod_clasificare,
        "cod_subunitate": cod_subunitate,
        "an_bugetar": an_bugetar,
        "anulat": {"$ne": True},
        "tip_operatie": {"$ne": "DIMINUARE"}
    }
    
    if exclude_angajament_id:
        match_filter["_id"] = {"$ne": ObjectId(exclude_angajament_id)}
    
    pipeline = [
        {"$match": match_filter},
        {"$group": {"_id": None, "total": {"$sum": "$suma_lei"}}}
    ]
    
    result = list(db.alop_angajamente_bugetare.aggregate(pipeline))
    total_angajat = Decimal(str(result[0]["total"])) if result else Decimal(0)
    
    return suma_aprobata - total_angajat


@router.get("/nomenclatoare/clasificari")
async def get_clasificari(
    current_user: dict = Depends(verify_token)
) -> List[Dict[str, Any]]:
    """Get clasificari for angajamente bugetare"""
    db = get_db(domain=current_user.get('domain'))
    try:
        clasificari = list(db.alop_clasificari.find().sort('cod', 1))
        result = []
        for item in clasificari:
            cod = _extract_first(item, CODE_FIELD_CANDIDATES) or ""
            denumire = _extract_first(item, NAME_FIELD_CANDIDATES) or ""
            label = f"{cod} - {denumire}".strip(" -")
            result.append({
                "value": str(item["_id"]),
                "label": label or cod or str(item["_id"]),
                "cod": cod,
                "denumire": denumire
            })
        return result
    except Exception as e:
        logger.error(subject="Error fetching clasificari", content=str(e), category="angajamente-bugetare")
        return []


@router.get("/nomenclatoare/subunitati")
async def get_subunitati(
    current_user: dict = Depends(verify_token)
) -> List[Dict[str, Any]]:
    """Get subunitati for angajamente bugetare"""
    db = get_db(domain=current_user.get('domain'))
    try:
        subunitati = list(db.alop_subunitati.find().sort('cod', 1))
        result = []
        for item in subunitati:
            cod = _extract_first(item, CODE_FIELD_CANDIDATES) or ""
            denumire = _extract_first(item, NAME_FIELD_CANDIDATES) or ""
            label = f"{cod} - {denumire}".strip(" -")
            result.append({
                "value": str(item["_id"]),
                "label": label or cod or str(item["_id"]),
                "cod": cod,
                "denumire": denumire
            })
        return result
    except Exception as e:
        logger.error(subject="Error fetching subunitati", content=str(e), category="angajamente-bugetare")
        return []


@router.get("/next-number")
async def get_next_numar_angajament(
    an_bugetar: Optional[int] = None,
    data_angajament: Optional[str] = None,
    current_user: dict = Depends(verify_token)
) -> Dict[str, Any]:
    """Return next available angajament number (without year in value)."""
    db = get_db(domain=current_user.get('domain'))
    try:
        year = an_bugetar
        if data_angajament:
            try:
                parsed = datetime.fromisoformat(data_angajament.replace('Z', '+00:00'))
                year = parsed.year
            except Exception:
                pass
        if not year:
            year = datetime.utcnow().year
        last = db.alop_angajamente_bugetare.find_one(
            {"an_bugetar": year},
            sort=[("creat_la", -1)]
        )
        last_num, last_width = _parse_number_with_width(last.get("numar_angajament") if last else None)
        start_num, start_width = _extract_config_start(db, "alop_angajamente_bugetare_start")
        width = max(last_width, start_width)
        if last_num is None:
            next_num = start_num if start_num is not None else 1
        else:
            if start_num is not None and last_num < start_num:
                next_num = start_num
            else:
                next_num = last_num + 1
        next_number = str(next_num).zfill(width)
        return {"next_number": next_number, "an_bugetar": year}
    except Exception as e:
        logger.error(subject="Error computing next number", content=str(e), category="angajamente-bugetare")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("")
async def get_angajamente_bugetare(
    page: int = Query(1, ge=1),
    limit: int = Query(50, ge=1, le=500),
    search: Optional[str] = None,
    an_bugetar: Optional[int] = None,
    stare: Optional[str] = None,
    cod_beneficiar: Optional[str] = None,
    cod_clasificare: Optional[str] = None,
    sort_by: str = "data_angajament",
    sort_order: str = "desc",
    user = Depends(verify_token)
):
    """
    Obține lista angajamentelor bugetare cu paginare și filtre
    """
    db = get_db(domain=user.get('domain'))
    
    try:
        # Construire query
        query_filter = {}
        
        if search:
            search_regex = {"$regex": search, "$options": "i"}
            query_filter["$or"] = [
                {"numar_angajament": search_regex},
                {"scop": search_regex},
                {"beneficiar.denumire": search_regex}
            ]
        
        if an_bugetar:
            query_filter["an_bugetar"] = an_bugetar
        
        if stare:
            query_filter["stare"] = stare
        
        if cod_beneficiar:
            query_filter["cod_beneficiar"] = cod_beneficiar
        
        if cod_clasificare:
            clasificari = list(db.alop_clasificari.find(
                {"$or": [
                    {"cod": {"$regex": f"^{cod_clasificare}", "$options": "i"}},
                    {"cod_clasificare": {"$regex": f"^{cod_clasificare}", "$options": "i"}}
                ]},
                {"_id": 1}
            ))
            ids = [c.get("_id") for c in clasificari if c.get("_id")]
            if ids:
                query_filter["clasificare_id"] = {"$in": ids}
            else:
                query_filter["clasificare_id"] = {"$in": []}
        
        # Count total
        total = db.alop_angajamente_bugetare.count_documents(query_filter)
        
        # Calculate skip
        skip = (page - 1) * limit
        
        # Sort direction
        sort_direction = -1 if sort_order == "desc" else 1
        
        # Query
        cursor = db.alop_angajamente_bugetare.find(query_filter).sort(sort_by, sort_direction).skip(skip).limit(limit)
        
        items = list(cursor)
        clasificari_ids = set()
        subunitati_ids = set()
        for item in items:
            clas_id = _normalize_object_id(item.get("clasificare_id"))
            sub_id = _normalize_object_id(item.get("subunitate_id"))
            if clas_id:
                clasificari_ids.add(clas_id)
                item["clasificare_id"] = clas_id
            if sub_id:
                subunitati_ids.add(sub_id)
                item["subunitate_id"] = sub_id
        clasificari_map = {}
        subunitati_map = {}
        if clasificari_ids:
            clasificari_docs = db.alop_clasificari.find({"_id": {"$in": list(clasificari_ids)}})
            for doc in clasificari_docs:
                cod = _extract_first(doc, CODE_FIELD_CANDIDATES)
                clasificari_map[str(doc["_id"])] = cod
        if subunitati_ids:
            subunitati_docs = db.alop_subunitati.find({"_id": {"$in": list(subunitati_ids)}})
            for doc in subunitati_docs:
                cod = _extract_first(doc, CODE_FIELD_CANDIDATES)
                subunitati_map[str(doc["_id"])] = cod

        result_items = []
        for item in items:
            item["_id"] = str(item["_id"])
            if item.get("clasificare_id"):
                item["clasificare_id"] = str(item["clasificare_id"])
                item["cod_clasificare"] = clasificari_map.get(item["clasificare_id"])
            if item.get("subunitate_id"):
                item["subunitate_id"] = str(item["subunitate_id"])
                item["cod_subunitate"] = subunitati_map.get(item["subunitate_id"])
            # Convert Decimal128 to float for JSON
            for key in ["suma_lei", "suma_valuta", "suma_totala", "suma_consumata", "suma_disponibila"]:
                if key in item and item[key]:
                    item[key] = float(str(item[key]))
            result_items.append(item)
        
        return {
            "items": result_items,
            "total": total,
            "page": page,
            "limit": limit,
            "pages": (total + limit - 1) // limit
        }
    
    except Exception as e:
        logger.error(subject="Error fetching angajamente bugetare", content=str(e), category="angajamente-bugetare")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/{angajament_id}")
async def get_angajament_bugetar(
    angajament_id: str,
    user = Depends(verify_token)
):
    """Obține detaliile unui angajament bugetar"""
    db = get_db(domain=user.get('domain'))
    
    try:
        angajament = db.alop_angajamente_bugetare.find_one({"_id": ObjectId(angajament_id)})
        
        if not angajament:
            raise HTTPException(status_code=404, detail="Angajamentul nu există")
        
        angajament["_id"] = str(angajament["_id"])
        if angajament.get("clasificare_id"):
            angajament["clasificare_id"] = str(angajament["clasificare_id"])
            try:
                doc = db.alop_clasificari.find_one({"_id": ObjectId(angajament["clasificare_id"])})
                angajament["cod_clasificare"] = _extract_first(doc or {}, CODE_FIELD_CANDIDATES)
            except Exception:
                angajament["cod_clasificare"] = angajament.get("cod_clasificare")
        if angajament.get("subunitate_id"):
            angajament["subunitate_id"] = str(angajament["subunitate_id"])
            try:
                doc = db.alop_subunitati.find_one({"_id": ObjectId(angajament["subunitate_id"])})
                angajament["cod_subunitate"] = _extract_first(doc or {}, CODE_FIELD_CANDIDATES)
            except Exception:
                angajament["cod_subunitate"] = angajament.get("cod_subunitate")
        
        # Convert Decimal128 to float
        for key in ["suma_lei", "suma_valuta", "suma_totala", "suma_consumata", "suma_disponibila"]:
            if key in angajament and angajament[key]:
                angajament[key] = float(str(angajament[key]))
        
        return angajament
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(subject=f"Error fetching angajament {angajament_id}", content=str(e), category="angajamente-bugetare")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("")
async def create_angajament_bugetar(
    data: Dict[str, Any],
    user = Depends(verify_token)
):
    """
    Creează un nou angajament bugetar
    
    Validări:
    - Suma nu poate depăși disponibilul din buget
    - Beneficiarul trebuie să existe
    - Cod clasificare valid
    """
    db = get_db(domain=user.get('domain'))
    
    try:
        # Extragere date
        clasificare_id = data.get("clasificare_id")
        subunitate_id = data.get("subunitate_id")
        cod_clasificare_input = data.get("cod_clasificare")
        cod_subunitate_input = data.get("cod_subunitate")
        suma_lei = Decimal(str(data.get("suma_lei", 0)))
        cod_beneficiar = data.get("cod_beneficiar")
        scop = data.get("scop")
        numar_angajament = data.get("numar_angajament")
        data_angajament = data.get("data_angajament")
        
        # Validări minime
        if not all([suma_lei > 0, numar_angajament, data_angajament]):
            raise HTTPException(status_code=400, detail="Câmpuri obligatorii lipsă")

        # Parsare data și an bugetar
        if isinstance(data_angajament, str):
            data_angajament = datetime.fromisoformat(data_angajament.replace('Z', '+00:00'))
        elif isinstance(data_angajament, date):
            data_angajament = datetime.combine(data_angajament, datetime.min.time())
        an_bugetar = data_angajament.year if data_angajament else data.get("an_bugetar")

        # Resolve clasificare și subunitate
        clas_oid, cod_clasificare, _ = _resolve_reference(
            db,
            "alop_clasificari",
            clasificare_id,
            cod_clasificare_input,
            CODE_FIELD_CANDIDATES
        )
        sub_oid, cod_subunitate, _ = _resolve_reference(
            db,
            "alop_subunitati",
            subunitate_id,
            cod_subunitate_input,
            CODE_FIELD_CANDIDATES
        )

        if not all([cod_clasificare, cod_subunitate, an_bugetar]):
            raise HTTPException(status_code=400, detail="Câmpuri obligatorii lipsă")

        # Verificare disponibil buget
        disponibil = await calculeaza_disponibil_buget(
            cod_clasificare,
            cod_subunitate,
            an_bugetar,
            db=db
        )
        
        if suma_lei > disponibil:
            raise HTTPException(
                status_code=400,
                detail=f"Suma ({suma_lei}) depășește disponibilul din buget ({disponibil})"
            )
        
        # Verificare beneficiar
        beneficiar_data = None
        if cod_beneficiar:
            beneficiar = db.alop_firme.find_one({"cod": cod_beneficiar})
            if not beneficiar:
                raise HTTPException(status_code=404, detail=f"Beneficiarul {cod_beneficiar} nu există")
            
            beneficiar_data = {
                "denumire": beneficiar.get("denumire", ""),
                "cif": beneficiar.get("cif", ""),
                "adresa": beneficiar.get("adresa_completa") or beneficiar.get("adresa", "")
            }
        
        # Verificare unicitate număr angajament
        existing = db.alop_angajamente_bugetare.find_one({
            "numar_angajament": numar_angajament,
            "an_bugetar": an_bugetar
        })
        
        if existing:
            raise HTTPException(status_code=400, detail=f"Angajamentul {numar_angajament} există deja pentru anul {an_bugetar}")
        
        # Generare număr CFP Propunere
        cfpp_filter = {
            "an_bugetar": an_bugetar,
            "cfp_propunere.numar": {"$exists": True}
        }
        if sub_oid:
            cfpp_filter["subunitate_id"] = sub_oid
        else:
            cfpp_filter["cod_subunitate"] = cod_subunitate

        ultimul_cfpp = db.alop_angajamente_bugetare.find_one(
            cfpp_filter,
            sort=[("cfp_propunere.numar", -1)]
        )
        
        if ultimul_cfpp and ultimul_cfpp.get("cfp_propunere", {}).get("numar"):
            try:
                ultimul_nr = int(ultimul_cfpp["cfp_propunere"]["numar"])
                numar_cfpp = str(ultimul_nr + 1).zfill(6)
            except:
                numar_cfpp = "000001"
        else:
            numar_cfpp = "000001"
        
        # Construire document
        document = {
            "numar_angajament": numar_angajament,
            "data_angajament": data_angajament,
            "an_bugetar": an_bugetar,
            
            "clasificare_id": clas_oid,
            "subunitate_id": sub_oid,
            "cod_clasificare_parsed": parseaza_cod_clasificare(cod_clasificare),
            
            "suma_lei": Decimal128(suma_lei),
            "suma_valuta": Decimal128(Decimal(str(data.get("suma_valuta", 0)))),
            "cod_moneda": data.get("cod_moneda", "LEI"),
            "suma_totala": Decimal128(Decimal(str(data.get("suma_totala", suma_lei)))),
            "suma_consumata": Decimal128(Decimal(0)),
            "suma_disponibila": Decimal128(Decimal(str(data.get("suma_totala", suma_lei)))),
            
            "cod_beneficiar": cod_beneficiar,
            "beneficiar": beneficiar_data,
            
            "cfp_propunere": {
                "numar": numar_cfpp,
                "data": data_angajament,
                "aprobat": True
            },
            "cfp_decont": None,
            
            "scop": scop,
            "tip_angajament": "MULTIANUAL" if data.get("este_multianual") else "NORMAL",
            "este_multianual": data.get("este_multianual", False),
            "cod_compartiment": data.get("cod_compartiment"),
            "cod_program": data.get("cod_program"),
            "delegat": data.get("delegat"),
            
            "stare": "ACTIV",
            "tip_operatie": "ANGAJARE",
            "anulat": False,
            "motiv_anulare": None,
            "data_anulare": None,

            "generated_docs": [],
            "signed_pdf_hash": None,
            "signed_pdf_filename": None,
            "signed_pdf_uploaded_at": None,
            "signed_pdf_uploaded_by": None,
            "aprobat": False,
            "aprobat_la": None,
            "aprobat_de": None,
            "aprobat_de_id": None,
            
            "angajamente_legale": [],
            
            "creat_la": datetime.utcnow(),
            "actualizat_la": datetime.utcnow(),
            "utilizator_creare": user.get("username", ""),
            "utilizator_actualizare": user.get("username", "")
        }
        
        result = db.alop_angajamente_bugetare.insert_one(document)
        
        logger.info(
            subject="Angajament bugetar creat",
            content=f"Nr: {numar_angajament}, Suma: {suma_lei}, User: {user.get('username')}",
            category="angajamente-bugetare"
        )
        
        return {
            "success": True,
            "id": str(result.inserted_id),
            "numar_angajament": numar_angajament,
            "numar_cfpp": numar_cfpp
        }
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(subject="Error creating angajament bugetar", content=str(e), category="angajamente-bugetare")
        raise HTTPException(status_code=500, detail=str(e))


@router.put("/{angajament_id}")
async def update_angajament_bugetar(
    angajament_id: str,
    data: Dict[str, Any],
    user = Depends(verify_token)
):
    """
    Modifică un angajament bugetar
    
    Validări:
    - Nu se poate reduce suma sub valoarea consumată
    - Suma nouă nu poate depăși disponibilul
    """
    db = get_db(domain=user.get('domain'))
    
    try:
        angajament = db.alop_angajamente_bugetare.find_one({"_id": ObjectId(angajament_id)})
        
        if not angajament:
            raise HTTPException(status_code=404, detail="Angajamentul nu există")
        
        if angajament["anulat"]:
            raise HTTPException(status_code=400, detail="Nu se poate modifica un angajament anulat")
        
        update_fields = {
            "actualizat_la": datetime.utcnow(),
            "utilizator_actualizare": user.get("username", "")
        }
        
        # Modificare sumă
        if "suma_lei" in data:
            suma_noua = Decimal(str(data["suma_lei"]))
            suma_consumata = Decimal(str(angajament["suma_consumata"]))
            
            if suma_noua < suma_consumata:
                raise HTTPException(
                    status_code=400,
                    detail=f"Suma nouă ({suma_noua}) nu poate fi mai mică decât suma consumată ({suma_consumata})"
                )
            
            # Verificare disponibil dacă se mărește
            suma_veche = Decimal(str(angajament["suma_lei"]))
            if suma_noua > suma_veche:
                cod_clasificare = angajament.get("cod_clasificare")
                cod_subunitate = angajament.get("cod_subunitate")
                if not cod_clasificare and angajament.get("clasificare_id"):
                    try:
                        doc = db.alop_clasificari.find_one({"_id": angajament["clasificare_id"]})
                        cod_clasificare = _extract_first(doc or {}, CODE_FIELD_CANDIDATES)
                    except Exception:
                        cod_clasificare = None
                if not cod_subunitate and angajament.get("subunitate_id"):
                    try:
                        doc = db.alop_subunitati.find_one({"_id": angajament["subunitate_id"]})
                        cod_subunitate = _extract_first(doc or {}, CODE_FIELD_CANDIDATES)
                    except Exception:
                        cod_subunitate = None

                disponibil = await calculeaza_disponibil_buget(
                    cod_clasificare,
                    cod_subunitate,
                    angajament["an_bugetar"],
                    exclude_angajament_id=angajament_id,
                    db=db
                )
                
                if suma_noua > disponibil:
                    raise HTTPException(
                        status_code=400,
                        detail=f"Suma nouă depășește disponibilul ({disponibil})"
                    )
            
            update_fields["suma_lei"] = Decimal128(suma_noua)
            update_fields["suma_totala"] = Decimal128(suma_noua)
            update_fields["suma_disponibila"] = Decimal128(suma_noua - suma_consumata)
        
        # Alte câmpuri
        if "scop" in data:
            update_fields["scop"] = data["scop"]
        
        if "cod_beneficiar" in data:
            beneficiar = db.alop_firme.find_one({"cod": data["cod_beneficiar"]})
            if beneficiar:
                update_fields["cod_beneficiar"] = data["cod_beneficiar"]
                update_fields["beneficiar"] = {
                    "denumire": beneficiar.get("denumire", ""),
                    "cif": beneficiar.get("cif", ""),
                    "adresa": beneficiar.get("adresa_completa") or beneficiar.get("adresa", "")
                }
        
        result = db.alop_angajamente_bugetare.update_one(
            {"_id": ObjectId(angajament_id)},
            {"$set": update_fields}
        )
        
        logger.info(
            subject="Angajament bugetar modificat",
            content=f"ID: {angajament_id}, User: {user.get('username')}",
            category="angajamente-bugetare"
        )
        
        return {"success": True, "modified_count": result.modified_count}
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(subject=f"Error updating angajament {angajament_id}", content=str(e), category="angajamente-bugetare")
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/{angajament_id}")
async def delete_angajament_bugetar(
    angajament_id: str,
    motiv: str = Query(..., description="Motivul anulării"),
    user = Depends(verify_token)
):
    """
    Anulează un angajament bugetar
    
    Validări:
    - Nu se poate anula dacă are angajamente legale active
    """
    db = get_db(domain=user.get('domain'))
    
    try:
        angajament = db.alop_angajamente_bugetare.find_one({"_id": ObjectId(angajament_id)})
        
        if not angajament:
            raise HTTPException(status_code=404, detail="Angajamentul nu există")
        
        if angajament["anulat"]:
            raise HTTPException(status_code=400, detail="Angajamentul este deja anulat")
        
        # Verificare angajamente legale (când vor fi implementate)
        # angajamente_legale_active = db.alop_angajamente_legale.count_documents({
        #     "angajament_bugetar_id": ObjectId(angajament_id),
        #     "anulat": {"$ne": True}
        # })
        # if angajamente_legale_active > 0:
        #     raise HTTPException(status_code=400, detail="Există angajamente legale active")
        
        result = db.alop_angajamente_bugetare.update_one(
            {"_id": ObjectId(angajament_id)},
            {
                "$set": {
                    "anulat": True,
                    "stare": "ANULAT",
                    "motiv_anulare": motiv,
                    "data_anulare": datetime.utcnow(),
                    "actualizat_la": datetime.utcnow(),
                    "utilizator_actualizare": user.get("username", "")
                }
            }
        )
        
        logger.info(
            subject="Angajament bugetar anulat",
            content=f"ID: {angajament_id}, Motiv: {motiv}, User: {user.get('username')}",
            category="angajamente-bugetare"
        )
        
        return {"success": True, "modified_count": result.modified_count}
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(subject=f"Error deleting angajament {angajament_id}", content=str(e), category="angajamente-bugetare")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/disponibil/buget")
async def get_disponibil_buget(
    an_bugetar: int = Query(...),
    cod_clasificare: Optional[str] = None,
    cod_subunitate: Optional[str] = None,
    clasificare_id: Optional[str] = None,
    subunitate_id: Optional[str] = None,
    user = Depends(verify_token)
):
    """Calculează disponibilul din buget pentru o clasificare"""
    db = get_db(domain=user.get('domain'))
    try:
        _, resolved_cod, _ = _resolve_reference(
            db,
            "alop_clasificari",
            clasificare_id,
            cod_clasificare,
            CODE_FIELD_CANDIDATES
        )
        _, resolved_sub, _ = _resolve_reference(
            db,
            "alop_subunitati",
            subunitate_id,
            cod_subunitate,
            CODE_FIELD_CANDIDATES
        )
        if not resolved_cod or not resolved_sub:
            raise HTTPException(status_code=400, detail="Clasificarea sau subunitatea lipsesc")

        disponibil = await calculeaza_disponibil_buget(
            resolved_cod,
            resolved_sub,
            an_bugetar,
            db=db
        )
        
        return {
            "cod_clasificare": resolved_cod,
            "cod_subunitate": resolved_sub,
            "an_bugetar": an_bugetar,
            "disponibil": float(disponibil)
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/export/excel")
async def export_angajamente_excel(
    an_bugetar: Optional[int] = None,
    stare: Optional[str] = None,
    search: Optional[str] = None,
    user = Depends(verify_token)
):
    """Export angajamente bugetare în format Excel"""
    from fastapi.responses import StreamingResponse
    import io
    
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")
    
    db = get_db(domain=user.get('domain'))
    
    try:
        # Construire query
        query_filter = {}
        
        if search:
            search_regex = {"$regex": search, "$options": "i"}
            query_filter["$or"] = [
                {"numar_angajament": search_regex},
                {"scop": search_regex},
                {"beneficiar.denumire": search_regex}
            ]
        
        if an_bugetar:
            query_filter["an_bugetar"] = an_bugetar
        
        if stare:
            query_filter["stare"] = stare
        
        # Query toate datele (fără paginare pentru export)
        cursor = db.alop_angajamente_bugetare.find(query_filter).sort("data_angajament", -1).limit(10000)
        items = list(cursor)
        clasificari_ids = set()
        subunitati_ids = set()
        for item in items:
            clas_id = _normalize_object_id(item.get("clasificare_id"))
            sub_id = _normalize_object_id(item.get("subunitate_id"))
            if clas_id:
                clasificari_ids.add(clas_id)
                item["clasificare_id"] = clas_id
            if sub_id:
                subunitati_ids.add(sub_id)
                item["subunitate_id"] = sub_id
        clasificari_map = {}
        subunitati_map = {}
        if clasificari_ids:
            for doc in db.alop_clasificari.find({"_id": {"$in": list(clasificari_ids)}}):
                clasificari_map[str(doc["_id"])] = _extract_first(doc, CODE_FIELD_CANDIDATES)
        if subunitati_ids:
            for doc in db.alop_subunitati.find({"_id": {"$in": list(subunitati_ids)}}):
                subunitati_map[str(doc["_id"])] = _extract_first(doc, CODE_FIELD_CANDIDATES)
        
        # Creare workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Angajamente Bugetare"
        
        # Header styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Headers
        headers = [
            "Nr. Angajament",
            "Data",
            "An Bugetar",
            "Cod Clasificare",
            "Beneficiar",
            "CIF",
            "Sumă Lei",
            "Sumă Consumată",
            "Disponibil",
            "Stare",
            "Tip",
            "Scop",
            "CFP Propunere"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Data rows
        row = 2
        for item in items:
            ws.cell(row=row, column=1, value=item.get("numar_angajament", ""))
            ws.cell(row=row, column=2, value=item.get("data_angajament").strftime("%d.%m.%Y") if item.get("data_angajament") else "")
            ws.cell(row=row, column=3, value=item.get("an_bugetar", ""))
            cod_clasificare = item.get("cod_clasificare")
            if not cod_clasificare and item.get("clasificare_id"):
                cod_clasificare = clasificari_map.get(str(item.get("clasificare_id")))
            ws.cell(row=row, column=4, value=cod_clasificare or "")
            
            beneficiar = item.get("beneficiar", {})
            ws.cell(row=row, column=5, value=beneficiar.get("denumire", "") if beneficiar else "")
            ws.cell(row=row, column=6, value=beneficiar.get("cif", "") if beneficiar else "")
            
            ws.cell(row=row, column=7, value=float(str(item.get("suma_lei", 0))))
            ws.cell(row=row, column=8, value=float(str(item.get("suma_consumata", 0))))
            ws.cell(row=row, column=9, value=float(str(item.get("suma_disponibila", 0))))
            ws.cell(row=row, column=10, value=item.get("stare", ""))
            ws.cell(row=row, column=11, value=item.get("tip_angajament", ""))
            ws.cell(row=row, column=12, value=item.get("scop", ""))
            
            cfp = item.get("cfp_propunere", {})
            ws.cell(row=row, column=13, value=cfp.get("numar", "") if cfp else "")
            
            row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Return as streaming response
        filename = f"angajamente_bugetare_{an_bugetar or 'toate'}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    
    except Exception as e:
        logger.error(subject="Error exporting angajamente to Excel", content=str(e), category="angajamente-bugetare")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/{angajament_id}/documents/generate")
async def generate_angajament_bugetar_document(
    angajament_id: str,
    request: AngajamentGenerateRequest,
    current_user: dict = Depends(verify_token)
):
    """Generate angajament bugetar document using DataFlows Docu."""
    db = get_db(domain=current_user.get('domain'))

    try:
        angajament = db.alop_angajamente_bugetare.find_one({"_id": ObjectId(angajament_id)})
        if not angajament:
            raise HTTPException(status_code=404, detail="Angajamentul nu există")

        template_code = request.template_code or "6X8DMFTLLM4D"
        template_name = request.template_name or "Angajament Bugetar"

        beneficiar = angajament.get("beneficiar") or {}
        cfp_propunere = angajament.get("cfp_propunere") or {}
        cod_clasificare = angajament.get("cod_clasificare")
        if not cod_clasificare and angajament.get("clasificare_id"):
            try:
                doc = db.alop_clasificari.find_one({"_id": ObjectId(str(angajament.get("clasificare_id")))})
                cod_clasificare = _extract_first(doc or {}, CODE_FIELD_CANDIDATES)
            except Exception:
                cod_clasificare = None
        cod_subunitate = angajament.get("cod_subunitate")
        if not cod_subunitate and angajament.get("subunitate_id"):
            try:
                doc = db.alop_subunitati.find_one({"_id": ObjectId(str(angajament.get("subunitate_id")))})
                cod_subunitate = _extract_first(doc or {}, CODE_FIELD_CANDIDATES)
            except Exception:
                cod_subunitate = None
        payload_data = {
            "data": {
                "id": str(angajament.get("_id")),
                "numar_angajament": angajament.get("numar_angajament"),
                "data_angajament": _format_date(angajament.get("data_angajament")),
                "an_bugetar": angajament.get("an_bugetar"),
                "cod_clasificare": cod_clasificare,
                "cod_subunitate": cod_subunitate,
                "cod_beneficiar": angajament.get("cod_beneficiar"),
                "beneficiar": beneficiar,
                "suma_lei": _decimal_to_float(angajament.get("suma_lei")),
                "suma_valuta": _decimal_to_float(angajament.get("suma_valuta")),
                "suma_totala": _decimal_to_float(angajament.get("suma_totala")),
                "suma_disponibila": _decimal_to_float(angajament.get("suma_disponibila")),
                "cod_moneda": angajament.get("cod_moneda", "LEI"),
                "scop": angajament.get("scop"),
                "tip_angajament": angajament.get("tip_angajament"),
                "cfp_propunere": {
                    "numar": cfp_propunere.get("numar"),
                    "data": _format_date(cfp_propunere.get("data")),
                    "aprobat": cfp_propunere.get("aprobat")
                }
            },
            "generated_at": datetime.utcnow().isoformat(),
            "generated_by": current_user.get("username", "unknown")
        }

        client = DataFlowsDocuClient()
        filename = f"angajament_bugetar_{angajament.get('numar_angajament', '')}_{angajament_id[:8]}"
        document_bytes = client.create_realtime_job(
            template_code=template_code,
            data=payload_data,
            format="pdf",
            filename=filename
        )

        if not document_bytes:
            raise HTTPException(status_code=500, detail="Failed to generate document")

        file_hash = save_document_file(document_bytes, f"{filename}.pdf")
        generated_at = datetime.utcnow()

        generated_docs = list(angajament.get("generated_docs", []))
        doc_id = str(ObjectId())
        updated_entry = {
            "id": doc_id,
            "template_code": template_code,
            "template_name": template_name,
            "file_hash": file_hash,
            "filename": f"{filename}.pdf",
            "generated_at": generated_at,
            "generated_by": current_user.get("username", "unknown")
        }

        replaced = False
        for idx, entry in enumerate(generated_docs):
            if entry.get("template_code") == template_code:
                updated_entry["id"] = entry.get("id", doc_id)
                generated_docs[idx] = updated_entry
                replaced = True
                break

        if not replaced:
            generated_docs.append(updated_entry)

        db.alop_angajamente_bugetare.update_one(
            {"_id": ObjectId(angajament_id)},
            {"$set": {"generated_docs": generated_docs, "actualizat_la": datetime.utcnow()}}
        )

        updated_entry["generated_at"] = generated_at.isoformat()

        return {"success": True, "document": updated_entry}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(subject="Error generating angajament bugetar document", content=str(e), category="angajamente-bugetare")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/{angajament_id}/documents/{doc_entry_id}/download")
async def download_angajament_bugetar_document(
    angajament_id: str,
    doc_entry_id: str,
    current_user: dict = Depends(verify_token)
):
    """Download generated angajament bugetar document."""
    db = get_db(domain=current_user.get('domain'))

    try:
        angajament = db.alop_angajamente_bugetare.find_one({"_id": ObjectId(angajament_id)})
        if not angajament:
            raise HTTPException(status_code=404, detail="Angajamentul nu există")

        docs = angajament.get("generated_docs", [])
        entry = next((doc for doc in docs if doc.get("id") == doc_entry_id), None)
        if not entry or not entry.get("file_hash"):
            raise HTTPException(status_code=404, detail="Document not found")

        file_path = get_file_path(entry.get("file_hash"))
        if not file_path or not os.path.exists(file_path):
            raise HTTPException(status_code=404, detail="File not found")

        filename = entry.get("filename", f"angajament_bugetar_{angajament_id}.pdf")
        return FileResponse(
            file_path,
            media_type="application/pdf",
            headers={"Content-Disposition": f'inline; filename=\"{filename}\"'}
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(subject="Error downloading angajament bugetar document", content=str(e), category="angajamente-bugetare")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/{angajament_id}/signed/upload")
async def upload_signed_angajament_bugetar(
    angajament_id: str,
    file: UploadFile = File(...),
    current_user: dict = Depends(verify_token)
):
    """Upload signed angajament bugetar PDF (economic department only)."""
    db = get_db(domain=current_user.get('domain'))

    try:
        angajament = db.alop_angajamente_bugetare.find_one({"_id": ObjectId(angajament_id)})
        if not angajament:
            raise HTTPException(status_code=404, detail="Angajamentul nu există")

        if not (is_admin_user(current_user) or _is_economic_department(current_user)):
            raise HTTPException(status_code=403, detail="Acces interzis pentru încărcare document semnat")

        if not file.filename or not file.filename.lower().endswith(".pdf"):
            raise HTTPException(status_code=400, detail="Doar fișiere PDF sunt permise")

        metadata = await save_upload_file(file)
        now = datetime.utcnow()

        db.alop_angajamente_bugetare.update_one(
            {"_id": ObjectId(angajament_id)},
            {"$set": {
                "signed_pdf_hash": metadata.get("hash"),
                "signed_pdf_filename": metadata.get("original_filename", file.filename),
                "signed_pdf_uploaded_at": now,
                "signed_pdf_uploaded_by": current_user.get("username", "unknown"),
                "aprobat": True,
                "aprobat_la": now,
                "aprobat_de": current_user.get("username", "unknown"),
                "aprobat_de_id": str(current_user.get("_id", "")),
                "actualizat_la": now
            }}
        )

        return {
            "success": True,
            "file_hash": metadata.get("hash"),
            "filename": metadata.get("original_filename", file.filename)
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(subject="Error uploading signed angajament bugetar", content=str(e), category="angajamente-bugetare")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/{angajament_id}/signed/download")
async def download_signed_angajament_bugetar(
    angajament_id: str,
    current_user: dict = Depends(verify_token)
):
    """Download signed angajament bugetar PDF."""
    db = get_db(domain=current_user.get('domain'))

    try:
        angajament = db.alop_angajamente_bugetare.find_one({"_id": ObjectId(angajament_id)})
        if not angajament:
            raise HTTPException(status_code=404, detail="Angajamentul nu există")

        file_hash = angajament.get("signed_pdf_hash")
        if not file_hash:
            raise HTTPException(status_code=404, detail="Signed document not found")

        file_path = get_file_path(file_hash)
        if not file_path or not os.path.exists(file_path):
            raise HTTPException(status_code=404, detail="File not found")

        filename = angajament.get("signed_pdf_filename", f"angajament_bugetar_{angajament_id}_signed.pdf")
        return FileResponse(
            file_path,
            media_type="application/pdf",
            headers={"Content-Disposition": f'inline; filename=\"{filename}\"'}
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(subject="Error downloading signed angajament bugetar", content=str(e), category="angajamente-bugetare")
        raise HTTPException(status_code=500, detail=str(e))
